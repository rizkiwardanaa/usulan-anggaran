import streamlit as st
import pandas as pd
import os
from io import BytesIO
from datetime import datetime
from sqlalchemy import create_engine
import streamlit.components.v1 as components

# --- KONEKSI KE CLOUD DATABASE ---
DB_URL = st.secrets["DB_URL"]
engine = create_engine(DB_URL, pool_size=10, max_overflow=20)

# =====================================================================
# FUNGSI DATABASE & HELPER (BUG WIPE DATA DIPERBAIKI)
# =====================================================================
@st.cache_data(ttl=300)
def load_table(table_name, default_cols):
    try:
        with engine.connect() as conn:
            df = pd.read_sql(f"SELECT * FROM {table_name}", conn)
            
        for col in default_cols:
            if col not in df.columns:
                if "Vol" in col or "Harga" in col or "Total" in col: df[col] = 1 if "Vol" in col else 0
                elif col == "Tahun": df[col] = str(datetime.now().year + 1)
                elif col == "Sumber_Dana": df[col] = "BOPTN"
                elif col == "Sub_Komponen" and table_name == "rab_m_akun": df[col] = "-"
                elif col == "Versi_RAB": df[col] = "Indikatif"
                elif col == "Is_Active": df[col] = 1
                else: df[col] = "-"
                
        if "Is_Active" in df.columns:
            df["Is_Active"] = pd.to_numeric(df["Is_Active"], errors='coerce').fillna(1).astype(int)
        return df
        
    except Exception as e:
        err_str = str(e).lower()
        if "does not exist" in err_str or "not found" in err_str or "relation" in err_str:
            df = pd.DataFrame(columns=default_cols)
            with engine.connect() as conn:
                df.to_sql(table_name, conn, if_exists="append", index=False)
            return df
        else:
            st.error(f"Koneksi ke tabel {table_name} terganggu. Error: {e}")
            return pd.DataFrame(columns=default_cols)

def save_table(df, table_name):
    with engine.connect() as conn:
        df.to_sql(table_name, conn, if_exists="replace", index=False)
    st.cache_data.clear()

def format_rupiah(x):
    try: return f"{float(x):,.0f}".replace(',', '.')
    except (ValueError, TypeError): return x

def split_kode(teks):
    s = str(teks).strip()
    if " - " in s:
        parts = s.split(" - ", 1)
        return parts[0].strip(), parts[1].strip()
    parts = s.split(" ", 1)
    if len(parts) == 2:
        first_part = parts[0].strip()
        if any(c.isdigit() for c in first_part) or len(first_part) <= 8 or "." in first_part:
            return first_part, parts[1].strip()
    if any(c.isdigit() for c in s) or len(s) <= 8 or "." in s:
        return s, ""
    return "", s

def get_vol_sat_combined(v1, s1, v2, s2):
    v1_str = str(v1).replace(".0", "") if pd.notna(v1) else "0"
    s1_str = str(s1).strip() if pd.notna(s1) else ""
    v2_str = str(v2).replace(".0", "") if pd.notna(v2) else "0"
    s2_str = str(s2).strip() if pd.notna(s2) else ""
    if s2_str in ["", "-"] or v2_str == "0" or v2_str == "":
        return f"{v1_str} {s1_str}"
    return f"{v1_str} {s1_str} x {v2_str} {s2_str}"

# =====================================================================
# GENERATOR CETAK RAB SATUAN (EXCEL & PDF)
# =====================================================================
def export_excel_rab(df_header, df_items, kegiatan_code_map):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "RAB Export"
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    
    font_bold = Font(bold=True); font_header = Font(bold=True, size=11); align_center = Alignment(horizontal="center", vertical="center")
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:E1')
    t_rab = df_header.get('Tahun', pd.Series(['2027'])).iloc[0]
    ws['A1'] = f"RINCIAN ANGGARAN BIAYA (RAB) FAKULTAS ILMU BUDAYA\nTAHUN ANGGARAN {t_rab}"
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['A1'].font = font_header
    ws.row_dimensions[1].height = 40

    keg_nama = df_header['Kegiatan'].iloc[0].title()
    keg_kode = kegiatan_code_map.get(df_header['Kegiatan'].iloc[0], "0000")

    meta_rows = [
        ("Kementerian/ Lembaga:", "(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI"), 
        ("Unit Eselon II/ Satker:", "(17) Dirjen Diktiristek / (677524) UNIVERSITAS MULAWARMAN"),
        ("Sumber Dana:", df_header.get('Sumber_Dana', pd.Series(['BOPTN'])).iloc[0]),
        ("Kegiatan:", f"{keg_kode} - {keg_nama}"), 
        ("Sasaran Kegiatan:", df_header['Sasaran'].iloc[0]), 
        ("Klasifikasi Rincian Output:", df_header['KRO'].iloc[0]),
        ("Volume:", df_header['Volume'].iloc[0]), ("Satuan Ukur:", df_header['Satuan'].iloc[0]), 
        ("Alokasi Dana (Total Belanja):", f"Rp. {df_items['Total_Biaya'].sum():,.0f}".replace(',','.'))
    ]
    rp = 2
    for label, val in meta_rows:
        ws.cell(row=rp, column=1, value=label).font = font_bold
        ws.cell(row=rp, column=2, value=val)
        ws.merge_cells(start_row=rp, start_column=2, end_row=rp, end_column=5)
        rp += 1

    rp += 1
    for col_idx, text in enumerate(["Kode", "Rincian Belanja", "Volume & Satuan", "Harga Satuan", "Jumlah Biaya"], start=1):
        cell = ws.cell(row=rp, column=col_idx, value=text); cell.font = font_bold; cell.alignment = align_center; cell.border = border_all
    rp += 1

    def print_row(kode, urai, vol, hrg, tot, is_bold=False):
        nonlocal rp
        ws.cell(row=rp, column=1, value=kode).border = border_all
        ws.cell(row=rp, column=2, value=urai).border = border_all
        ws.cell(row=rp, column=3, value=vol).border = border_all
        ws.cell(row=rp, column=4, value=hrg).border = border_all
        if hrg != "": ws.cell(row=rp, column=4).number_format = '#,##0'
        ws.cell(row=rp, column=5, value=tot).border = border_all; ws.cell(row=rp, column=5).number_format = '#,##0'
        if is_bold: 
            for col in range(1,6): ws.cell(row=rp, column=col).font = Font(bold=True)
        rp += 1

    total_seluruh = df_items["Total_Biaya"].sum()
    for head_col, indent in [('RO', ""), ('Komponen', "  "), ('Sub_Komponen', "    ")]:
        if df_header[head_col].iloc[0] and str(df_header[head_col].iloc[0]).strip() not in ["", "-", "Tidak Ada Sub-Komponen"]:
            k_val, u_val = split_kode(df_header[head_col].iloc[0])
            print_row(k_val, f"{indent}{u_val}", "", "", total_seluruh, True)

    print_row(keg_kode, f"      {keg_nama}", "", "", total_seluruh, True)

    for akun, group in df_items.groupby("Akun_Belanja"):
        k_ak, u_ak = split_kode(akun)
        print_row(k_ak, f"        {u_ak}", "", "", group['Total_Biaya'].sum(), True)
        for _, r in group.iterrows():
            v_sat = get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2'])
            print_row("", f"          - {r['Uraian']}", v_sat, r['Harga_Satuan'], r['Total_Biaya'])
            
    rp += 2
    bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    try: 
        tobj = datetime.strptime(df_header['Tgl_Cetak'].iloc[0], "%Y-%m-%d")
        tgl_str = f"Samarinda, {tobj.day} {bulan_indo[tobj.month-1]} {tobj.year}"
    except: tgl_str = f"Samarinda, {df_header['Tgl_Cetak'].iloc[0]}"
    
    ws.cell(row=rp, column=4, value=tgl_str)
    ws.cell(row=rp+1, column=4, value=df_header['Jabatan'].iloc[0])
    ws.cell(row=rp+5, column=4, value=df_header['Nama_Pejabat'].iloc[0]).font = Font(underline="single", bold=True)
    ws.cell(row=rp+6, column=4, value=f"NIP. {df_header['NIP_Pejabat'].iloc[0]}")

    output = BytesIO(); wb.save(output)
    return output.getvalue()

def export_pdf_rab(df_header, df_items, orientasi, kegiatan_code_map):
    total_seluruh = df_items["Total_Biaya"].sum()
    t_rab = df_header.get('Tahun', pd.Series(['2027'])).iloc[0]
    s_dana = df_header.get('Sumber_Dana', pd.Series(['BOPTN'])).iloc[0]
    keg_nama_full = df_header['Kegiatan'].iloc[0].title()
    keg_kode_full = kegiatan_code_map.get(df_header['Kegiatan'].iloc[0], "0000")
    
    try: 
        tobj = datetime.strptime(df_header['Tgl_Cetak'].iloc[0], "%Y-%m-%d")
        bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        tgl_str = f"Samarinda, {tobj.day} {bulan_indo[tobj.month-1]} {tobj.year}"
    except: tgl_str = f"Samarinda, {df_header['Tgl_Cetak'].iloc[0]}"
    
    page_rule = "A4 landscape" if orientasi == "Landscape" else "A4 portrait"
    
    html = f"""
    <!DOCTYPE html>
    <html><head><meta charset="utf-8">
    <style>
        @page {{ size: {page_rule}; margin: 10mm; }}
        body {{ font-family: 'Arial', sans-serif; font-size: 8.5pt; line-height: 1.2; color: #000; }}
        .judul {{ text-align: center; font-weight: bold; font-size: 11pt; margin-bottom: 15px; }}
        .tabel-meta td {{ padding: 1px 3px; font-size: 8.5pt; }}
        .tabel-utama {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 8pt; }}
        .tabel-utama th, .tabel-utama td {{ border: 1px solid black; padding: 4px; }}
        .tabel-utama th {{ background-color: #d9d9d9; text-align: center; font-weight: bold; }}
        .text-right {{ text-align: right; }} .text-center {{ text-align: center; }} .bold {{ font-weight: bold; }}
        .ttd-box {{ width: 220px; float: right; text-align: left; margin-top: 20px; margin-right: 15px; page-break-inside: avoid; }}
        .kro-row {{ background-color: #d9e1f2; }} .ro-row {{ background-color: #e9edf4; }}
        .komp-row {{ background-color: #fff2cc; }} .sub-row {{ background-color: #fce4d6; }}
        .keg-row {{ background-color: #e2efda; }} .akun-row {{ background-color: #ffffff; }}
    </style></head><body>
    <div class="judul">RINCIAN ANGGARAN BIAYA (RAB) FAKULTAS ILMU BUDAYA<br>TAHUN ANGGARAN {t_rab}</div>
    <table class="tabel-meta">
        <tr><td class="bold">Kementerian/ Lembaga</td><td>:</td><td>(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI</td></tr>
        <tr><td class="bold">Unit Eselon II/ Satker</td><td>:</td><td>(17) Dirjen Diktiristek / (677524) UNIVERSITAS MULAWARMAN</td></tr>
        <tr><td class="bold">Sumber Dana</td><td>:</td><td>{s_dana}</td></tr>
        <tr><td class="bold">Kegiatan</td><td>:</td><td>{keg_kode_full} - {keg_nama_full}</td></tr>
        <tr><td class="bold">Sasaran Kegiatan</td><td>:</td><td>{df_header['Sasaran'].iloc[0]}</td></tr>
        <tr><td class="bold">Klasifikasi Rincian Output</td><td>:</td><td>{df_header['KRO'].iloc[0]}</td></tr>
        <tr><td class="bold">Volume</td><td>:</td><td>{df_header['Volume'].iloc[0]}</td></tr>
        <tr><td class="bold">Satuan Ukur</td><td>:</td><td>{df_header['Satuan'].iloc[0]}</td></tr>
        <tr><td class="bold">Alokasi Dana (Total Belanja)</td><td>:</td><td>Rp. {format_rupiah(total_seluruh)}</td></tr>
    </table>
    <table class="tabel-utama">
        <tr><th>Kode</th><th>Rincian Belanja</th><th>Volume & Satuan</th><th>Harga Satuan</th><th>Jumlah Biaya</th></tr>
    """
    for head_col, indent, cls_row in [('RO', "", "ro-row"), ('Komponen', "  ", "komp-row"), ('Sub_Komponen', "    ", "sub-row")]:
        if df_header[head_col].iloc[0] and str(df_header[head_col].iloc[0]).strip() not in ["", "-", "Tidak Ada Sub-Komponen"]:
            k, u = split_kode(df_header[head_col].iloc[0])
            html += f"<tr class='{cls_row} bold'><td>{k}</td><td>{indent}{u}</td><td></td><td></td><td class='right'>{format_rupiah(total_seluruh)}</td></tr>"
    
    html += f"<tr class='keg-row bold'><td>{keg_kode_full}</td><td style='padding-left:15px;'>{keg_nama_full}</td><td></td><td></td><td class='right'>{format_rupiah(total_seluruh)}</td></tr>"

    for akun, group_akun in df_items.groupby("Akun_Belanja"):
        k_ak, u_ak = split_kode(akun)
        html += f"<tr class='akun-row bold'><td>{k_ak}</td><td style='padding-left:30px;'>{u_ak}</td><td></td><td></td><td class='right'>{format_rupiah(group_akun['Total_Biaya'].sum())}</td></tr>"
        for _, r in group_akun.iterrows():
            v_sat_str = get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2'])
            html += f"<tr><td></td><td style='padding-left:45px;'>- {r['Uraian']}</td><td class='center'>{v_sat_str}</td><td class='right'>{format_rupiah(r['Harga_Satuan'])}</td><td class='right'>{format_rupiah(r['Total_Biaya'])}</td></tr>"
    
    html += f"""</table>
    <div class="ttd-box">
        {tgl_str}<br>{df_header['Jabatan'].iloc[0]}<br><br><br><br><br>
        <b><u>{df_header['Nama_Pejabat'].iloc[0]}</u></b><br>NIP. {df_header['NIP_Pejabat'].iloc[0]}
    </div>
    </body></html>"""
    return html

# =====================================================================
# GENERATOR MATRIK PERUBAHAN & RKAKL
# =====================================================================
def generate_matrik_html(df_matrik, v_sebelum, v_menjadi, keg_map):
    if df_matrik.empty: return "<h3>Tidak ada data untuk dibandingkan pada versi tersebut.</h3>"
    
    html = f"""
    <!DOCTYPE html>
    <html><head><meta charset="utf-8">
    <style>
        @page {{ size: A4 landscape; margin: 15mm; }}
        body {{ font-family: 'Arial', sans-serif; font-size: 7.5pt; line-height: 1.2; color: #000; }}
        .center {{ text-align: center; }} .right {{ text-align: right; }} .bold {{ font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 7.5pt; }}
        th, td {{ border: 1px solid black; padding: 4px; vertical-align: top; }}
        th {{ background-color: #d9d9d9; text-align: center; font-weight: bold; }}
        .kro-row {{ background-color: #d9e1f2; }} .ro-row {{ background-color: #e9edf4; }}
        .komp-row {{ background-color: #fff2cc; }} .sub-row {{ background-color: #fce4d6; }}
        .keg-row {{ background-color: #e2efda; }}
    </style></head><body>
    <h3 class="center" style="margin-bottom:2px;">MATRIK PERUBAHAN RENCANA KERJA DAN ANGGARAN</h3>
    <h4 class="center" style="margin-top:0px; margin-bottom:20px;">Versi {v_sebelum} menjadi {v_menjadi}</h4>
    <table>
        <tr>
            <th width="7%" rowspan="2">KODE</th>
            <th width="28%" rowspan="2">URAIAN PROGRAM / KEGIATAN /<br>KOMPONEN / AKUN / DETAIL</th>
            <th width="20%" colspan="3">PAGU SEMULA ({v_sebelum})</th>
            <th width="20%" colspan="3">PAGU MENJADI ({v_menjadi})</th>
            <th width="10%" rowspan="2">BERTAMBAH /<br>(BERKURANG)</th>
            <th width="5%" rowspan="2">DANA</th>
        </tr>
        <tr>
            <th>VOL</th><th>HARGA</th><th>JUMLAH</th>
            <th>VOL</th><th>HARGA</th><th>JUMLAH</th>
        </tr>
    """
    tot_s_global, tot_m_global, tot_sel_global = 0, 0, 0
    
    for kro, g_kro in df_matrik.groupby('KRO'):
        k_kro, n_kro = split_kode(kro)
        sd = g_kro['Sumber_Dana'].iloc[0]
        t_s = g_kro['Tot_s'].sum(); t_m = g_kro['Tot_m'].sum(); selisih = g_kro['Selisih'].sum()
        tot_s_global += t_s; tot_m_global += t_m; tot_sel_global += selisih
        html += f"<tr class='kro-row bold'><td>{k_kro}</td><td>{n_kro}</td><td></td><td></td><td class='right'>{format_rupiah(t_s)}</td><td></td><td></td><td class='right'>{format_rupiah(t_m)}</td><td class='right'>{format_rupiah(selisih)}</td><td class='center'>{sd}</td></tr>"
        
        for ro, g_ro in g_kro.groupby('RO'):
            k_ro, n_ro = split_kode(ro)
            t_s = g_ro['Tot_s'].sum(); t_m = g_ro['Tot_m'].sum(); selisih = g_ro['Selisih'].sum()
            html += f"<tr class='ro-row bold'><td>{k_ro}</td><td>{n_ro}</td><td></td><td></td><td class='right'>{format_rupiah(t_s)}</td><td></td><td></td><td class='right'>{format_rupiah(t_m)}</td><td class='right'>{format_rupiah(selisih)}</td><td></td></tr>"
            
            for komp, g_komp in g_ro.groupby('Komponen'):
                k_komp, n_komp = split_kode(komp)
                t_s = g_komp['Tot_s'].sum(); t_m = g_komp['Tot_m'].sum(); selisih = g_komp['Selisih'].sum()
                html += f"<tr class='komp-row bold'><td>{k_komp}</td><td>{n_komp}</td><td></td><td></td><td class='right'>{format_rupiah(t_s)}</td><td></td><td></td><td class='right'>{format_rupiah(t_m)}</td><td class='right'>{format_rupiah(selisih)}</td><td></td></tr>"
                
                for sub, g_sub in g_komp.groupby('Sub_Komponen'):
                    if sub and sub != "-":
                        k_sub, n_sub = split_kode(sub)
                        t_s = g_sub['Tot_s'].sum(); t_m = g_sub['Tot_m'].sum(); selisih = g_sub['Selisih'].sum()
                        html += f"<tr class='sub-row bold'><td>{k_sub}</td><td>{n_sub}</td><td></td><td></td><td class='right'>{format_rupiah(t_s)}</td><td></td><td></td><td class='right'>{format_rupiah(t_m)}</td><td class='right'>{format_rupiah(selisih)}</td><td></td></tr>"
                    
                    for keg, g_keg in g_sub.groupby('Kegiatan'):
                        k_keg = keg_map.get(keg, "0000"); n_keg = keg.title()
                        t_s = g_keg['Tot_s'].sum(); t_m = g_keg['Tot_m'].sum(); selisih = g_keg['Selisih'].sum()
                        html += f"<tr class='keg-row bold'><td>{k_keg}</td><td style='padding-left:10px;'>{n_keg}</td><td></td><td></td><td class='right'>{format_rupiah(t_s)}</td><td></td><td></td><td class='right'>{format_rupiah(t_m)}</td><td class='right'>{format_rupiah(selisih)}</td><td></td></tr>"
                        
                        for akun, g_akun in g_keg.groupby('Akun_Belanja'):
                            k_ak, n_ak = split_kode(akun)
                            t_s = g_akun['Tot_s'].sum(); t_m = g_akun['Tot_m'].sum(); selisih = g_akun['Selisih'].sum()
                            html += f"<tr class='bold'><td>{k_ak}</td><td style='padding-left:20px;'>{n_ak}</td><td></td><td></td><td class='right'>{format_rupiah(t_s)}</td><td></td><td></td><td class='right'>{format_rupiah(t_m)}</td><td class='right'>{format_rupiah(selisih)}</td><td></td></tr>"
                            
                            for _, det in g_akun.iterrows():
                                v_s = get_vol_sat_combined(det['V1_s'], det['S1_s'], det['V2_s'], det['S2_s']) if det['Tot_s'] > 0 else "-"
                                v_m = get_vol_sat_combined(det['V1_m'], det['S1_m'], det['V2_m'], det['S2_m']) if det['Tot_m'] > 0 else "-"
                                h_s = format_rupiah(det['Hrg_s']) if det['Tot_s'] > 0 else "-"
                                h_m = format_rupiah(det['Hrg_m']) if det['Tot_m'] > 0 else "-"
                                html += f"<tr><td></td><td style='padding-left:30px;'>- {det['Uraian']}</td><td class='center'>{v_s}</td><td class='right'>{h_s}</td><td class='right'>{format_rupiah(det['Tot_s'])}</td><td class='center'>{v_m}</td><td class='right'>{h_m}</td><td class='right'>{format_rupiah(det['Tot_m'])}</td><td class='right'>{format_rupiah(det['Selisih'])}</td><td></td></tr>"

    html += f"""<tr class='bold' style='background-color:#d9d9d9;'><td colspan='2' class='right'>TOTAL GLOBAL</td><td></td><td></td><td class='right'>Rp {format_rupiah(tot_s_global)}</td><td></td><td></td><td class='right'>Rp {format_rupiah(tot_m_global)}</td><td class='right'>Rp {format_rupiah(tot_sel_global)}</td><td></td></tr></table></body></html>"""
    return html

def generate_rkakl_html(df_utama, df_detail, kegiatan_code_map):
    if df_utama.empty: return "<h3>Belum ada data RAB aktif.</h3>"
    html = """
    <!DOCTYPE html>
    <html><head><meta charset="utf-8">
    <style>
        @page { size: A4 portrait; margin: 10mm; }
        body { font-family: 'Arial', sans-serif; font-size: 7.5pt; line-height: 1.2; color: #000; }
        .center { text-align: center; } .right { text-align: right; } .bold { font-weight: bold; }
        table { width: 100%; border-collapse: collapse; margin-top: 5px; font-size: 7.5pt; }
        th, td { border: 1px solid black; padding: 4px 5px; vertical-align: top; }
        th { background-color: #d9d9d9; text-align: center; font-weight: bold; }
        .kro-row { background-color: #d9e1f2; } .ro-row { background-color: #e9edf4; }
        .komp-row { background-color: #fff2cc; } .sub-row { background-color: #fce4d6; }
        .keg-row { background-color: #e2efda; }
    </style></head><body>
    <h3 class="center" style="margin-bottom:2px;">LAPORAN RENCANA KERJA DAN ANGGARAN (RKAKL)</h3>
    <h4 class="center" style="margin-top:0px; margin-bottom:15px;">FAKULTAS ILMU BUDAYA - UNIVERSITAS MULAWARMAN</h4>
    <table>
        <tr>
            <th width="10%">KODE</th>
            <th width="40%">PROGRAM / KEGIATAN / OUTPUT / SUBOUTPUT /<br>KOMPONEN / SUBKOMP / JUDUL KEGIATAN / AKUN / DETIL</th>
            <th width="12%">VOL</th><th width="14%">HARGA SATUAN</th><th width="14%">JUMLAH BIAYA</th><th width="10%">DANA</th>
        </tr>
    """
    total_semua = 0
    for kro, g_kro in df_utama.groupby('KRO'):
        k_kro, n_kro = split_kode(kro)
        s_dana = g_kro['Sumber_Dana'].iloc[0]
        ids_kro = g_kro['ID_RAB'].tolist()
        tot_kro = df_detail[df_detail['ID_RAB'].isin(ids_kro)]['Total_Biaya'].sum()
        total_semua += tot_kro
        html += f"<tr class='kro-row bold'><td>{k_kro}</td><td>{n_kro}</td><td></td><td></td><td class='right'>{format_rupiah(tot_kro)}</td><td class='center'>{s_dana}</td></tr>"
        for ro, g_ro in g_kro.groupby('RO'):
            k_ro, n_ro = split_kode(ro)
            ids_ro = g_ro['ID_RAB'].tolist()
            tot_ro = df_detail[df_detail['ID_RAB'].isin(ids_ro)]['Total_Biaya'].sum()
            html += f"<tr class='ro-row bold'><td>{k_ro}</td><td>{n_ro}</td><td></td><td></td><td class='right'>{format_rupiah(tot_ro)}</td><td></td></tr>"
            for komp, g_komp in g_ro.groupby('Komponen'):
                k_komp, n_komp = split_kode(komp)
                ids_komp = g_komp['ID_RAB'].tolist()
                tot_komp = df_detail[df_detail['ID_RAB'].isin(ids_komp)]['Total_Biaya'].sum()
                html += f"<tr class='komp-row bold'><td>{k_komp}</td><td>{n_komp}</td><td></td><td></td><td class='right'>{format_rupiah(tot_komp)}</td><td></td></tr>"
                for sub, g_sub in g_komp.groupby('Sub_Komponen'):
                    if sub and sub != "-":
                        k_sub, n_sub = split_kode(sub)
                        ids_sub = g_sub['ID_RAB'].tolist()
                        tot_sub = df_detail[df_detail['ID_RAB'].isin(ids_sub)]['Total_Biaya'].sum()
                        html += f"<tr class='sub-row bold'><td>{k_sub}</td><td>{n_sub}</td><td></td><td></td><td class='right'>{format_rupiah(tot_sub)}</td><td></td></tr>"
                    for keg, g_keg in g_sub.groupby('Kegiatan'):
                        keg_code = kegiatan_code_map.get(keg, "0000"); keg_title = keg.title() 
                        ids_keg = g_keg['ID_RAB'].tolist()
                        tot_keg = df_detail[df_detail['ID_RAB'].isin(ids_keg)]['Total_Biaya'].sum()
                        html += f"<tr class='keg-row bold'><td>{keg_code}</td><td style='padding-left:10px;'>{keg_title}</td><td></td><td></td><td class='right'>{format_rupiah(tot_keg)}</td><td></td></tr>"
                        det_keg = df_detail[df_detail['ID_RAB'].isin(ids_keg)]
                        for akun, g_akun in det_keg.groupby('Akun_Belanja'):
                            k_akun, n_akun = split_kode(akun)
                            tot_akun = g_akun['Total_Biaya'].sum()
                            html += f"<tr class='bold'><td>{k_akun}</td><td style='padding-left:20px;'>{n_akun}</td><td></td><td></td><td class='right'>{format_rupiah(tot_akun)}</td><td></td></tr>"
                            for _, det in g_akun.iterrows():
                                v_sat = get_vol_sat_combined(det['Vol_1'], det['Sat_1'], det['Vol_2'], det['Sat_2'])
                                html += f"<tr><td></td><td style='padding-left:30px;'>- {det['Uraian']}</td><td class='center'>{v_sat}</td><td class='right'>{format_rupiah(det['Harga_Satuan'])}</td><td class='right'>{format_rupiah(det['Total_Biaya'])}</td><td></td></tr>"

    html += f"<tr class='bold' style='background-color:#d9d9d9;'><td colspan='4' class='right'>TOTAL SELURUH ANGGARAN (RKAKL AKTIF)</td><td class='right'>Rp {format_rupiah(total_semua)}</td><td></td></tr></table></body></html>"
    return html

# =====================================================================
# MODUL UTAMA MANAJEMEN HALAMAN
# =====================================================================
def show_page():
    # 1. Load Data
    df_m_kro = load_table("rab_m_kro", ["KRO", "Sumber_Dana"])
    df_m_ro = load_table("rab_m_ro", ["KRO", "RO", "Sumber_Dana"])
    df_m_komp = load_table("rab_m_komp", ["RO", "Komponen", "Sumber_Dana"])
    df_m_subkomp = load_table("rab_m_subkomp", ["Komponen", "Sub_Komponen", "Sumber_Dana"])
    df_m_akun = load_table("rab_m_akun", ["Sub_Komponen", "Account_Code", "Account_Name", "Sumber_Dana"]) 
    df_m_pejabat = load_table("rab_m_pejabat", ["Jabatan", "Nama", "NIP"])

    df_rab_utama = load_table("rab_utama", ["ID_RAB", "Tanggal", "Tahun", "Tgl_Cetak", "Sumber_Dana", "KRO", "RO", "Komponen", "Sub_Komponen", "Kegiatan", "Sasaran", "Volume", "Satuan", "Alokasi", "Jabatan", "Nama_Pejabat", "NIP_Pejabat", "Versi_RAB", "Is_Active"])
    df_rab_detail = load_table("rab_detail", ["ID_RAB", "Akun_Belanja", "Uraian", "Vol_1", "Sat_1", "Vol_2", "Sat_2", "Harga_Satuan", "Total_Biaya"])

    unique_kegiatans = sorted(df_rab_utama['Kegiatan'].unique()) if not df_rab_utama.empty else []
    kegiatan_code_map = {keg: f"{i+1:04d}" for i, keg in enumerate(unique_kegiatans)}

    if 'edit_rab_id' not in st.session_state: st.session_state.edit_rab_id = None

    st.title("📄 Pengolah Dokumen RAB Universitas")
    st.caption("Sistem Manajemen & Generator RAB Berjenjang dengan Pemisahan Kode Otomatis, Sumber Dana & Matrik Versi Anggaran.")

    list_tahun = sorted(df_rab_utama['Tahun'].unique().tolist() + [str(datetime.now().year + 1)], reverse=True)
    list_tahun = list(dict.fromkeys(list_tahun)) 
    tahun_aktif = st.sidebar.selectbox("📅 Pilih Tahun Anggaran Aktif:", list_tahun)

    tab_master, tab_buat, tab_daftar, tab_rekap, tab_matrik = st.tabs(["🗂️ Master", "📝 Buat / Edit RAB", "📂 Arsip & Versi", "📊 RKAKL Aktif", "⚖️ Matrik Perubahan"])

    # -----------------------------------------------------------------
    # TAB 1: MASTER DATABASE
    # -----------------------------------------------------------------
    with tab_master:
        st.info("Input Master Data secara manual atau gunakan file Excel untuk backup/restore massal.")
        with st.expander("💾 Import & Export Data Master (Excel)", expanded=False):
            c_eks, c_imp = st.columns(2)
            with c_eks:
                st.markdown("**1. Export Data Master**")
                output_master = BytesIO()
                with pd.ExcelWriter(output_master, engine='openpyxl') as writer:
                    df_m_kro.to_excel(writer, index=False, sheet_name='KRO')
                    df_m_ro.to_excel(writer, index=False, sheet_name='RO')
                    df_m_komp.to_excel(writer, index=False, sheet_name='Komponen')
                    df_m_subkomp.to_excel(writer, index=False, sheet_name='Sub_Komponen')
                    df_m_akun.to_excel(writer, index=False, sheet_name='Akun')
                    df_m_pejabat.to_excel(writer, index=False, sheet_name='Pejabat')
                st.download_button(label="📥 Download Backup Master (.xlsx)", data=output_master.getvalue(), file_name=f"Backup_Master_RAB_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
            
            with c_imp:
                st.markdown("**2. Import Data Master**")
                file_master = st.file_uploader("Upload File Backup Excel", type=['xlsx'])
                if st.button("🚀 Jalankan Import", type="primary"):
                    if file_master is not None:
                        try:
                            xls_master = pd.read_excel(file_master, sheet_name=None)
                            if 'KRO' in xls_master: save_table(xls_master['KRO'], "rab_m_kro")
                            if 'RO' in xls_master: save_table(xls_master['RO'], "rab_m_ro")
                            if 'Komponen' in xls_master: save_table(xls_master['Komponen'], "rab_m_komp")
                            if 'Sub_Komponen' in xls_master: save_table(xls_master['Sub_Komponen'], "rab_m_subkomp")
                            if 'Akun' in xls_master: save_table(xls_master['Akun'], "rab_m_akun")
                            if 'Pejabat' in xls_master: save_table(xls_master['Pejabat'], "rab_m_pejabat")
                            st.success("Data Master berhasil di-import!"); st.rerun()
                        except Exception as e: st.error(f"Gagal memproses file: {e}")

        sumber_master = st.radio("Pilih Kategori Master yang Ingin Diedit:", ["BOPTN", "PNBP"], horizontal=True)
        st.markdown("---")
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            st.markdown(f"**1. Master KRO ({sumber_master})**")
            df_kro_f = df_m_kro[df_m_kro['Sumber_Dana'] == sumber_master].copy()
            edit_kro = st.data_editor(df_kro_f[["KRO"]], num_rows="dynamic", use_container_width=True, hide_index=True, key="me_kro")
            if st.button("💾 Simpan KRO"): 
                edit_kro['Sumber_Dana'] = sumber_master
                df_sisa = df_m_kro[df_m_kro['Sumber_Dana'] != sumber_master]
                save_table(pd.concat([df_sisa, edit_kro.dropna(subset=["KRO"])]), "rab_m_kro"); st.rerun()
                
            st.markdown(f"**3. Master Komponen ({sumber_master})**")
            df_komp_f = df_m_komp[df_m_komp['Sumber_Dana'] == sumber_master].copy()
            edit_komp = st.data_editor(df_komp_f[["RO", "Komponen"]], num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"RO": st.column_config.SelectboxColumn(options=df_m_ro[df_m_ro['Sumber_Dana'] == sumber_master]["RO"].tolist())}, key="me_komp")
            if st.button("💾 Simpan Komponen"): 
                edit_komp['Sumber_Dana'] = sumber_master
                save_table(pd.concat([df_m_komp[df_m_komp['Sumber_Dana'] != sumber_master], edit_komp.dropna(subset=["Komponen"])]), "rab_m_komp"); st.rerun()

            st.markdown(f"**5. Master Akun Belanja ({sumber_master})**")
            df_akun_f = df_m_akun[df_m_akun['Sumber_Dana'] == sumber_master].copy()
            list_sub = df_m_subkomp[df_m_subkomp['Sumber_Dana'] == sumber_master]["Sub_Komponen"].tolist()
            edit_akun = st.data_editor(df_akun_f[["Sub_Komponen", "Account_Code", "Account_Name"]], num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Sub_Komponen": st.column_config.SelectboxColumn(options=list_sub if list_sub else ["-"])}, key="me_akun")
            if st.button("💾 Simpan Akun Belanja"): 
                edit_akun['Sumber_Dana'] = sumber_master
                save_table(pd.concat([df_m_akun[df_m_akun['Sumber_Dana'] != sumber_master], edit_akun.dropna(subset=["Account_Code", "Sub_Komponen"])]), "rab_m_akun"); st.rerun()

        with col_m2:
            st.markdown(f"**2. Master RO ({sumber_master})**")
            df_ro_f = df_m_ro[df_m_ro['Sumber_Dana'] == sumber_master].copy()
            list_kro = df_m_kro[df_m_kro['Sumber_Dana'] == sumber_master]["KRO"].tolist()
            edit_ro = st.data_editor(df_ro_f[["KRO", "RO"]], num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"KRO": st.column_config.SelectboxColumn(options=list_kro)}, key="me_ro")
            if st.button("💾 Simpan RO"): 
                edit_ro['Sumber_Dana'] = sumber_master
                save_table(pd.concat([df_m_ro[df_m_ro['Sumber_Dana'] != sumber_master], edit_ro.dropna(subset=["RO"])]), "rab_m_ro"); st.rerun()
            
            st.markdown(f"**4. Master Sub-Komponen ({sumber_master})**")
            df_sub_f = df_m_subkomp[df_m_subkomp['Sumber_Dana'] == sumber_master].copy()
            list_komp = df_m_komp[df_m_komp['Sumber_Dana'] == sumber_master]["Komponen"].tolist()
            edit_subkomp = st.data_editor(df_sub_f[["Komponen", "Sub_Komponen"]], num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Komponen": st.column_config.SelectboxColumn(options=list_komp)}, key="me_subkomp")
            if st.button("💾 Simpan Sub-Komponen"): 
                edit_subkomp['Sumber_Dana'] = sumber_master
                save_table(pd.concat([df_m_subkomp[df_m_subkomp['Sumber_Dana'] != sumber_master], edit_subkomp.dropna(subset=["Sub_Komponen"])]), "rab_m_subkomp"); st.rerun()

            st.markdown("**6. Master Pejabat**")
            edit_pejabat = st.data_editor(df_m_pejabat, num_rows="dynamic", use_container_width=True, hide_index=True, key="me_pej")
            if st.button("💾 Simpan Pejabat"): save_table(edit_pejabat.dropna(how='all'), "rab_m_pejabat"); st.rerun()

    # -----------------------------------------------------------------
    # TAB 2: BUAT / EDIT RAB BARU
    # -----------------------------------------------------------------
    with tab_buat:
        is_edit_mode = st.session_state.edit_rab_id is not None
        df_edit_head = pd.DataFrame()
        df_edit_det = pd.DataFrame()
        
        if is_edit_mode:
            st.info("✏️ **MODE REVISI:** Anda sedang mengedit/merevisi kegiatan. Data lama akan dimuat ke formulir.")
            if st.button("❌ Batal Edit (Kembali Buat Baru)"):
                st.session_state.edit_rab_id = None; st.rerun()
                
            df_edit_head = df_rab_utama[df_rab_utama['ID_RAB'] == st.session_state.edit_rab_id]
            df_edit_det = df_rab_detail[df_rab_detail['ID_RAB'] == st.session_state.edit_rab_id]
        
        def_sumber = df_edit_head['Sumber_Dana'].iloc[0] if not df_edit_head.empty else "BOPTN"
        def_kro = df_edit_head['KRO'].iloc[0] if not df_edit_head.empty else None
        def_ro = df_edit_head['RO'].iloc[0] if not df_edit_head.empty else None
        def_komp = df_edit_head['Komponen'].iloc[0] if not df_edit_head.empty else None
        def_subkomp = df_edit_head['Sub_Komponen'].iloc[0] if not df_edit_head.empty else None
        def_keg = df_edit_head['Kegiatan'].iloc[0] if not df_edit_head.empty else ""
        def_sasaran = df_edit_head['Sasaran'].iloc[0] if not df_edit_head.empty else ""
        def_vol = df_edit_head['Volume'].iloc[0] if not df_edit_head.empty else 1
        def_sat = df_edit_head['Satuan'].iloc[0] if not df_edit_head.empty else ""
        def_versi = df_edit_head['Versi_RAB'].iloc[0] if not df_edit_head.empty else "Indikatif"

        sumber_buat = st.radio("Pilih Sumber Dana RAB:", ["BOPTN", "PNBP"], index=["BOPTN", "PNBP"].index(def_sumber), horizontal=True, key="rb_buat")
        st.markdown("---")
        
        if df_m_kro.empty or df_m_ro.empty or df_m_komp.empty or df_m_akun.empty:
            st.warning("⚠️ Master Database masih kosong!")
        else:
            with st.container(border=True):
                st.subheader("1. Klasifikasi Output RAB")
                col_c1, col_c2 = st.columns(2)
                opsi_kro = df_m_kro[df_m_kro['Sumber_Dana'] == sumber_buat]["KRO"].tolist()
                idx_kro = opsi_kro.index(def_kro) if def_kro in opsi_kro else 0
                pilih_kro = col_c1.selectbox("Pilih KRO", opsi_kro if opsi_kro else ["-"], index=idx_kro)
                
                opsi_ro = df_m_ro[(df_m_ro['Sumber_Dana'] == sumber_buat) & (df_m_ro['KRO'] == pilih_kro)]["RO"].tolist()
                idx_ro = opsi_ro.index(def_ro) if def_ro in opsi_ro else 0
                pilih_ro = col_c2.selectbox("Pilih RO", opsi_ro if opsi_ro else ["-"], index=idx_ro)
                
                col_c3, col_c4 = st.columns(2)
                opsi_komp = df_m_komp[(df_m_komp['Sumber_Dana'] == sumber_buat) & (df_m_komp['RO'] == pilih_ro)]["Komponen"].tolist()
                idx_komp = opsi_komp.index(def_komp) if def_komp in opsi_komp else 0
                pilih_komp = col_c3.selectbox("Pilih Komponen", opsi_komp if opsi_komp else ["-"], index=idx_komp)
                
                opsi_subkomp = df_m_subkomp[(df_m_subkomp['Sumber_Dana'] == sumber_buat) & (df_m_subkomp['Komponen'] == pilih_komp)]["Sub_Komponen"].tolist()
                idx_subkomp = opsi_subkomp.index(def_subkomp) if def_subkomp in opsi_subkomp else 0
                pilih_subkomp = col_c4.selectbox("Pilih Sub-Komponen", opsi_subkomp if opsi_subkomp else ["-"], index=idx_subkomp)

            with st.container(border=True):
                st.subheader("2. Informasi Utama Kegiatan")
                col_u1, col_u2 = st.columns(2)
                rab_kegiatan = col_u1.text_input("Nama Kegiatan", value=def_keg, placeholder="Contoh: Pengadaan Peralatan Podcast")
                
                if not def_sasaran:
                    _, kro_narasi = split_kode(pilih_kro) if pilih_kro else ("", "")
                    def_sasaran = f"Peningkatan {kro_narasi.strip('() ')}" if kro_narasi else ""
                
                rab_sasaran = col_u2.text_input("Sasaran Kegiatan", value=def_sasaran)
                rab_vol = col_u1.number_input("Volume Target", value=int(def_vol), min_value=1)
                rab_satuan = col_u2.text_input("Satuan Ukur", value=def_sat, placeholder="Contoh: Layanan / Bulan")
                
                rab_tahun = col_u1.text_input("Tahun Anggaran", value=tahun_aktif, disabled=True)
                list_versi = ["Indikatif", "Definitif", "Revisi 1", "Revisi 2", "Revisi 3", "Revisi 4"]
                idx_versi = list_versi.index(def_versi) if def_versi in list_versi else 0
                rab_versi = col_u2.selectbox("Versi Anggaran (Periode)", list_versi, index=idx_versi)

            with st.container(border=True):
                st.subheader("3. Rincian Belanja")
                df_akun_f = df_m_akun[(df_m_akun['Sumber_Dana'] == sumber_buat) & (df_m_akun['Sub_Komponen'] == pilih_subkomp)]
                opsi_akun = [f"{row['Account_Code']} - {row['Account_Name']}" for _, row in df_akun_f.iterrows()]
                if not opsi_akun: opsi_akun = ["- Tidak ada akun terpetakan -"]
                
                if is_edit_mode and not df_edit_det.empty:
                    df_det_edit = df_edit_det.rename(columns={"Akun_Belanja":"Akun Belanja", "Uraian":"Uraian Belanja", "Vol_1":"Vol 1", "Sat_1":"Sat 1", "Vol_2":"Vol 2", "Sat_2":"Sat 2", "Harga_Satuan":"Harga Satuan"})
                    df_det_edit = df_det_edit[["Akun Belanja", "Uraian Belanja", "Vol 1", "Sat 1", "Vol 2", "Sat 2", "Harga Satuan"]]
                else:
                    df_det_edit = pd.DataFrame([{"Akun Belanja": opsi_akun[0], "Uraian Belanja": "", "Vol 1": 1, "Sat 1": "Unit", "Vol 2": 1, "Sat 2": "-", "Harga Satuan": 0}])
                
                df_input_detail = st.data_editor(
                    df_det_edit, num_rows="dynamic", use_container_width=True, hide_index=True, key="grid_buat_rab",
                    column_config={
                        "Akun Belanja": st.column_config.SelectboxColumn("Akun Belanja", options=opsi_akun, required=True),
                        "Uraian Belanja": st.column_config.TextColumn("Detail / Uraian", required=True),
                        "Vol 1": st.column_config.NumberColumn("Vol 1", min_value=1, required=True),
                        "Sat 1": st.column_config.TextColumn("Sat 1", required=True),
                        "Vol 2": st.column_config.NumberColumn("Vol 2", min_value=0),
                        "Sat 2": st.column_config.TextColumn("Sat 2 (Biarkan '-' jika tak ada)"),
                        "Harga Satuan": st.column_config.NumberColumn("Harga Satuan (Rp)", min_value=0, required=True)
                    }
                )

                df_input_detail["Vol_1_Num"] = pd.to_numeric(df_input_detail["Vol 1"]).fillna(1)
                df_input_detail["Vol_2_Num"] = pd.to_numeric(df_input_detail["Vol 2"]).fillna(1)
                df_input_detail.loc[df_input_detail["Vol_2_Num"] == 0, "Vol_2_Num"] = 1
                df_input_detail["Harga_Num"] = pd.to_numeric(df_input_detail["Harga Satuan"]).fillna(0)
                total_rab_live = (df_input_detail["Vol_1_Num"] * df_input_detail["Vol_2_Num"] * df_input_detail["Harga_Num"]).sum()
                
                c_pagu1, c_pagu2 = st.columns(2)
                is_pagu_locked = c_pagu1.checkbox("🔒 Aktifkan Kunci Pagu Maksimal (Opsional)")
                batas_pagu = c_pagu2.number_input("Batas Pagu (Rp)", min_value=0, value=int(total_rab_live)) if is_pagu_locked else 0
                
                st.markdown("#### 💰 Akumulasi Anggaran Alokasi Dana")
                if is_pagu_locked and total_rab_live > batas_pagu:
                    st.error(f"Total Anggaran (Rp {format_rupiah(total_rab_live)}) MELEBIHI Batas Pagu (Rp {format_rupiah(batas_pagu)})!")
                else:
                    st.metric(f"Total Alokasi Dana ({sumber_buat})", f"Rp {format_rupiah(total_rab_live)}")

            with st.container(border=True):
                st.subheader("4. Pengesahan & Simpan")
                col_p1, col_p2 = st.columns(2)
                opsi_pejabat = {idx: f"{row['Jabatan']} - {row['Nama']}" for idx, row in df_m_pejabat.iterrows()}
                pilih_pejabat = col_p1.selectbox("Pilih Pejabat Penandatangan", options=list(opsi_pejabat.keys()), format_func=lambda x: opsi_pejabat[x]) if opsi_pejabat else None
                tgl_cetak = col_p2.date_input("Tanggal Dokumen Cetak")
                
                if st.button("💾 Simpan RAB", type="primary"):
                    valid_detail = df_input_detail[df_input_detail["Uraian Belanja"].str.strip() != ""].copy()
                    if is_pagu_locked and total_rab_live > batas_pagu:
                        st.error("Gagal! Total rincian melebihi batas Pagu yang Anda kunci.")
                    elif not rab_kegiatan or valid_detail.empty or pilih_pejabat is None:
                        st.error("Gagal! Pastikan Nama Kegiatan, Rincian, dan Pejabat sudah terisi.")
                    else:
                        if is_edit_mode and def_versi == rab_versi:
                            df_rab_utama = df_rab_utama[df_rab_utama["ID_RAB"] != st.session_state.edit_rab_id]
                            df_rab_detail = df_rab_detail[df_rab_detail["ID_RAB"] != st.session_state.edit_rab_id]
                            
                        id_rab_baru = f"RAB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        dt_pjb = df_m_pejabat.loc[pilih_pejabat]
                        
                        # Set default active: if this version is currently the active version, keep it active
                        active_vs = df_rab_utama[(df_rab_utama['Tahun'] == tahun_aktif) & (df_rab_utama['Is_Active'] == 1)]['Versi_RAB'].unique()
                        is_act = 1 if len(active_vs) == 0 or rab_versi in active_vs else 0

                        new_utama = pd.DataFrame([{
                            "ID_RAB": id_rab_baru, "Tanggal": datetime.now().strftime('%Y-%m-%d %H:%M'), "Tahun": tahun_aktif, "Tgl_Cetak": str(tgl_cetak),
                            "Sumber_Dana": sumber_buat, "KRO": pilih_kro, "RO": pilih_ro, "Komponen": pilih_komp, "Sub_Komponen": pilih_subkomp,
                            "Kegiatan": rab_kegiatan.strip(), "Sasaran": rab_sasaran, "Volume": rab_vol, "Satuan": rab_satuan, "Alokasi": total_rab_live,
                            "Jabatan": dt_pjb['Jabatan'], "Nama_Pejabat": dt_pjb['Nama'], "NIP_Pejabat": dt_pjb['NIP'],
                            "Versi_RAB": rab_versi, "Is_Active": is_act
                        }])
                        df_rab_utama = pd.concat([df_rab_utama, new_utama], ignore_index=True)
                        save_table(df_rab_utama, "rab_utama")
                        
                        valid_detail["ID_RAB"] = id_rab_baru
                        valid_detail["Total_Biaya"] = valid_detail["Vol_1_Num"] * valid_detail["Vol_2_Num"] * valid_detail["Harga_Num"]
                        valid_detail.rename(columns={"Akun Belanja": "Akun_Belanja", "Uraian Belanja": "Uraian", "Vol 1":"Vol_1", "Sat 1":"Sat_1", "Vol 2":"Vol_2", "Sat 2":"Sat_2", "Harga Satuan": "Harga_Satuan"}, inplace=True)
                        
                        df_rab_detail = pd.concat([df_rab_detail, valid_detail[["ID_RAB", "Akun_Belanja", "Uraian", "Vol_1", "Sat_1", "Vol_2", "Sat_2", "Harga_Satuan", "Total_Biaya"]]], ignore_index=True)
                        save_table(df_rab_detail, "rab_detail")
                        
                        st.session_state.edit_rab_id = None
                        st.success(f"✅ RAB '{rab_kegiatan.title()}' Versi '{rab_versi}' Tersimpan!"); st.rerun()

    # -----------------------------------------------------------------
    # TAB 3: ARSIP & MANAJEMEN VERSI RAB (DENGAN FIX BUG COPY MASSAL & CETAK)
    # -----------------------------------------------------------------
    with tab_daftar:
        df_utama_thn = df_rab_utama[df_rab_utama['Tahun'] == tahun_aktif]
        if df_utama_thn.empty: 
            st.info(f"Belum ada dokumen RAB untuk Tahun {tahun_aktif}.")
        else:
            st.subheader("📂 Arsip & Manajemen Versi")
            col_a1, col_a2 = st.columns(2)
            versi_list_aktif = sorted(df_utama_thn['Versi_RAB'].unique())
            pilih_v_arsip = col_a1.selectbox("1. Pilih Versi (Untuk Difilter):", versi_list_aktif)
            
            df_v_terpilih = df_utama_thn[df_utama_thn['Versi_RAB'] == pilih_v_arsip]
            
            # PANEL STATUS VERSI GLOBAL
            is_v_active = 1 if 1 in df_v_terpilih['Is_Active'].values else 0
            if is_v_active == 1:
                st.success(f"✅ **STATUS VERSI: AKTIF (FINAL ACUAN)**. Seluruh kegiatan pada versi '{pilih_v_arsip}' ditarik ke dalam Rekapitulasi Global RKAKL.")
            else:
                st.warning(f"🗄️ **STATUS VERSI: ARSIP REVISI (TIDAK AKTIF)**. Versi ini disimpan sebagai rekaman sejarah anggaran.")
                if st.button(f"🔄 Jadikan Seluruh Kegiatan di Versi '{pilih_v_arsip}' Sebagai Acuan Aktif", type="primary"):
                    df_rab_utama.loc[df_rab_utama['Tahun'] == tahun_aktif, 'Is_Active'] = 0
                    df_rab_utama.loc[(df_rab_utama['Tahun'] == tahun_aktif) & (df_rab_utama['Versi_RAB'] == pilih_v_arsip), 'Is_Active'] = 1
                    save_table(df_rab_utama, "rab_utama")
                    st.toast(f"Versi {pilih_v_arsip} Berhasil Diaktifkan!")
                    st.rerun()

            st.markdown("---")

            keg_list_aktif = sorted(df_v_terpilih['Kegiatan'].unique())
            pilih_keg_arsip = col_a2.selectbox(f"2. Pilih Kegiatan (Dalam Versi {pilih_v_arsip}):", keg_list_aktif, format_func=lambda x: x.title())
            
            with st.expander(f"📋 Duplikasi Seluruh Kegiatan Versi '{pilih_v_arsip}' ke Versi Lain"):
                st.write("Salin seluruh kegiatan pada versi ini ke versi baru sekaligus. Sangat cocok digunakan sebelum membuat revisi RKAKL.")
                target_versi = st.selectbox("Pilih Target Versi Baru:", ["Indikatif", "Definitif", "Revisi 1", "Revisi 2", "Revisi 3", "Revisi 4"])
                if st.button(f"🚀 Salin Semua ke '{target_versi}'", type="primary"):
                    # Nonaktifkan versi lama di tahun yang sama
                    df_rab_utama.loc[df_rab_utama['Tahun'] == tahun_aktif, 'Is_Active'] = 0
                    for i, (_, row_keg) in enumerate(df_v_terpilih.iterrows()):
                        new_row = row_keg.copy()
                        new_id = f"RAB-{datetime.now().strftime('%Y%m%d%H%M%S%f')}-{i}"
                        new_row['ID_RAB'] = new_id
                        new_row['Versi_RAB'] = target_versi
                        new_row['Is_Active'] = 1
                        df_rab_utama = pd.concat([df_rab_utama, pd.DataFrame([new_row])], ignore_index=True)
                        
                        old_id = row_keg['ID_RAB']
                        det_keg_lama = df_rab_detail[df_rab_detail['ID_RAB'] == old_id].copy()
                        det_keg_lama['ID_RAB'] = new_id
                        df_rab_detail = pd.concat([df_rab_detail, det_keg_lama], ignore_index=True)
                        
                    save_table(df_rab_utama, "rab_utama"); save_table(df_rab_detail, "rab_detail")
                    st.success(f"Berhasil menduplikasi ke {target_versi} dan versi tersebut langsung diaktifkan!"); st.rerun()

            if pilih_keg_arsip:
                head_terpilih = df_v_terpilih[df_v_terpilih['Kegiatan'] == pilih_keg_arsip]
                id_rab_aktif = head_terpilih['ID_RAB'].iloc[0]
                detail_terpilih = df_rab_detail[df_rab_detail["ID_RAB"] == id_rab_aktif]
                
                if st.button("✏️ Edit Kegiatan Ini", use_container_width=True, type="secondary"):
                    st.session_state.edit_rab_id = id_rab_aktif; st.rerun()

                st.markdown("---")
                df_view = detail_terpilih.copy()
                df_view['Kode Akun'] = df_view['Akun_Belanja'].apply(lambda x: split_kode(x)[0])
                df_view['Nama Akun Belanja'] = df_view['Akun_Belanja'].apply(lambda x: split_kode(x)[1])
                df_view['Volume & Satuan'] = df_view.apply(lambda r: get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2']), axis=1)
                
                keg_code_view = kegiatan_code_map.get(pilih_keg_arsip, "0000")
                s_dana = head_terpilih.get('Sumber_Dana', pd.Series(['BOPTN'])).iloc[0]
                st.markdown(f"**Identitas Kegiatan:** {keg_code_view} - {pilih_keg_arsip.title()}")
                st.markdown(f"**Klasifikasi Dokumen:** {head_terpilih['KRO'].iloc[0]} ➔ {head_terpilih['RO'].iloc[0]}")
                st.markdown(f"**Total Alokasi Anggaran ({s_dana}):** Rp {format_rupiah(detail_terpilih['Total_Biaya'].sum())}")
                st.dataframe(df_view[["Kode Akun", "Nama Akun Belanja", "Uraian", "Volume & Satuan", "Harga_Satuan", "Total_Biaya"]].style.format({"Harga_Satuan": format_rupiah, "Total_Biaya": format_rupiah}), hide_index=True, use_container_width=True)

                st.markdown("#### 🖨️ Cetak Dokumen Satuan")
                col_x1, col_x2, col_x3 = st.columns([1, 1, 2])
                with col_x1:
                    st.download_button("📥 Download Excel Resmi", data=export_excel_rab(head_terpilih, detail_terpilih, kegiatan_code_map), file_name=f"RAB_{pilih_keg_arsip}_{pilih_v_arsip}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                with col_x2:
                    st.download_button("📑 Download PDF (Web)", data=export_pdf_rab(head_terpilih, detail_terpilih, "Portrait", kegiatan_code_map).encode('utf-8'), file_name=f"RAB_{pilih_keg_arsip}_{pilih_v_arsip}.html", mime="text/html", use_container_width=True)
                with col_x3:
                    if st.button("🗑️ Hapus Dokumen Ini", type="secondary", use_container_width=True):
                        df_rab_utama = df_rab_utama[df_rab_utama["ID_RAB"] != id_rab_aktif]
                        df_rab_detail = df_rab_detail[df_rab_detail["ID_RAB"] != id_rab_aktif]
                        save_table(df_rab_utama, "rab_utama"); save_table(df_rab_detail, "rab_detail"); st.rerun()

    # -----------------------------------------------------------------
    # TAB 4: REKAPITULASI RKAKL AKTIF
    # -----------------------------------------------------------------
    with tab_rekap:
        st.subheader(f"📊 Buku Rekapitulasi (RKAKL) Aktif Tahun {tahun_aktif}")
        df_aktif = df_rab_utama[(df_rab_utama['Is_Active'] == 1) & (df_rab_utama['Tahun'] == tahun_aktif)]
        if df_aktif.empty:
            st.info(f"Belum ada RAB aktif untuk tahun {tahun_aktif}.")
        else:
            df_det_aktif = df_rab_detail[df_rab_detail['ID_RAB'].isin(df_aktif['ID_RAB'])]
            html_rkakl = generate_rkakl_html(df_aktif, df_det_aktif, kegiatan_code_map)
            with st.container(border=True): components.html(html_rkakl, height=600, scrolling=True)
            st.download_button("📥 Cetak Buku Rekap RKAKL (.html)", data=html_rkakl.encode('utf-8'), file_name=f"RKAKL_FIB_{tahun_aktif}_{datetime.now().strftime('%Y%m%d')}.html", mime="text/html", type="primary")

    # -----------------------------------------------------------------
    # TAB 5: MATRIK PERUBAHAN
    # -----------------------------------------------------------------
    with tab_matrik:
        st.subheader("⚖️ Matrik Perbandingan Revisi Anggaran")
        df_thn = df_rab_utama[df_rab_utama['Tahun'] == tahun_aktif]
        if df_thn.empty:
            st.warning("Belum ada data untuk dibandingkan.")
        else:
            list_all_versions = sorted(df_thn['Versi_RAB'].unique())
            col_v1, col_v2 = st.columns(2)
            v1_def = list_all_versions[0] if len(list_all_versions) > 0 else None
            v2_def = list_all_versions[1] if len(list_all_versions) > 1 else v1_def
            pilih_v1 = col_v1.selectbox("Pilih Versi Semula (Sebelum):", list_all_versions, index=list_all_versions.index(v1_def) if v1_def else 0)
            pilih_v2 = col_v2.selectbox("Pilih Versi Menjadi (Sesudah):", list_all_versions, index=list_all_versions.index(v2_def) if v2_def else 0)
            
            if st.button("🔍 Generate Matrik Perbandingan", type="primary"):
                df_u1 = df_thn[df_thn['Versi_RAB'] == pilih_v1]
                df_d1 = df_rab_detail[df_rab_detail['ID_RAB'].isin(df_u1['ID_RAB'])]
                df_m1 = pd.merge(df_d1, df_u1, on='ID_RAB') if not df_u1.empty else pd.DataFrame()
                
                df_u2 = df_thn[df_thn['Versi_RAB'] == pilih_v2]
                df_d2 = df_rab_detail[df_rab_detail['ID_RAB'].isin(df_u2['ID_RAB'])]
                df_m2 = pd.merge(df_d2, df_u2, on='ID_RAB') if not df_u2.empty else pd.DataFrame()
                
                keys = ['Sumber_Dana', 'KRO', 'RO', 'Komponen', 'Sub_Komponen', 'Kegiatan', 'Akun_Belanja', 'Uraian']
                agg_dict = {'Vol_1': 'first', 'Sat_1': 'first', 'Vol_2': 'first', 'Sat_2': 'first', 'Harga_Satuan': 'first', 'Total_Biaya': 'sum'}
                
                if not df_m1.empty: df_m1 = df_m1.groupby(keys).agg(agg_dict).reset_index()
                else: df_m1 = pd.DataFrame(columns=keys + list(agg_dict.keys()))
                    
                if not df_m2.empty: df_m2 = df_m2.groupby(keys).agg(agg_dict).reset_index()
                else: df_m2 = pd.DataFrame(columns=keys + list(agg_dict.keys()))
                
                df_m1.rename(columns={'Vol_1':'V1_s', 'Sat_1':'S1_s', 'Vol_2':'V2_s', 'Sat_2':'S2_s', 'Harga_Satuan':'Hrg_s', 'Total_Biaya':'Tot_s'}, inplace=True)
                df_m2.rename(columns={'Vol_1':'V1_m', 'Sat_1':'S1_m', 'Vol_2':'V2_m', 'Sat_2':'S2_m', 'Harga_Satuan':'Hrg_m', 'Total_Biaya':'Tot_m'}, inplace=True)
                
                df_matrik = pd.merge(df_m1, df_m2, on=keys, how='outer')
                num_cols = ['V1_s', 'V2_s', 'Hrg_s', 'Tot_s', 'V1_m', 'V2_m', 'Hrg_m', 'Tot_m']
                str_cols = ['S1_s', 'S2_s', 'S1_m', 'S2_m']
                
                for c in num_cols: df_matrik[c] = df_matrik.get(c, pd.Series(0)).fillna(0)
                for c in str_cols: df_matrik[c] = df_matrik.get(c, pd.Series("")).fillna("")
                df_matrik['Selisih'] = df_matrik['Tot_m'] - df_matrik['Tot_s']
                
                if df_matrik.empty: st.info("Tidak ada data rincian pada kedua versi yang dipilih.")
                else:
                    html_matrik = generate_matrik_html(df_matrik, pilih_v1, pilih_v2, kegiatan_code_map)
                    with st.container(border=True): components.html(html_matrik, height=600, scrolling=True)
                    st.download_button("📥 Cetak Matrik Perubahan (.html)", data=html_matrik.encode('utf-8'), file_name=f"Matrik_{tahun_aktif}_{pilih_v1}_vs_{pilih_v2}.html", mime="text/html")
