import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from utils import load_table, get_available_years, split_kode, get_vol_sat_combined, format_rupiah, format_tgl_indo, update_rab_tahun, log_audit

st.title("📂 Arsip & Manajemen Versi RAB")

# --- LAZY LOADING ---
list_tahun = get_available_years()
tahun_aktif = st.sidebar.selectbox("📅 Pilih Tahun Anggaran Aktif:", list_tahun)

df_rab_utama = load_table("rab_utama", ["ID_RAB", "Tanggal", "Tahun", "Tgl_Cetak", "Sumber_Dana", "KRO", "RO", "Komponen", "Sub_Komponen", "Kegiatan", "Sasaran", "Volume", "Satuan", "Alokasi", "Jabatan", "Nama_Pejabat", "NIP_Pejabat", "Versi_RAB", "Is_Active", "Catatan"], f"WHERE \"Tahun\" = '{tahun_aktif}'")

if not df_rab_utama.empty:
    ids = tuple(df_rab_utama['ID_RAB'].tolist())
    where_det = f"WHERE \"ID_RAB\" = '{ids[0]}'" if len(ids) == 1 else f"WHERE \"ID_RAB\" IN {ids}"
    df_rab_detail = load_table("rab_detail", ["ID_RAB", "Akun_Belanja", "Uraian", "Vol_1", "Sat_1", "Vol_2", "Sat_2", "Harga_Satuan", "Total_Biaya"], where_det)
else:
    df_rab_detail = pd.DataFrame(columns=["ID_RAB", "Akun_Belanja", "Uraian", "Vol_1", "Sat_1", "Vol_2", "Sat_2", "Harga_Satuan", "Total_Biaya"])

unique_kegiatans = sorted(df_rab_utama['Kegiatan'].unique()) if not df_rab_utama.empty else []
kegiatan_code_map = {keg: f"{i+1:04d}" for i, keg in enumerate(unique_kegiatans)}

if 'edit_rab_id' not in st.session_state: st.session_state.edit_rab_id = None

# --- FUNGSI CETAK ---
def export_excel_rab(df_header, df_items, kegiatan_code_map, tampilkan_paraf=False):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "RAB Export"
    
    ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25; ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 16
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = 1; ws.page_setup.fitToWidth = 1
    
    font_bold = Font(bold=True); font_header = Font(bold=True, size=11); align_center = Alignment(horizontal="center", vertical="center")
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:E1')
    t_rab = df_header.get('Tahun', pd.Series(['2027'])).iloc[0]
    ws['A1'] = f"RINCIAN ANGGARAN BIAYA (RAB) FAKULTAS ILMU BUDAYA\nTAHUN ANGGARAN {t_rab}"
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['A1'].font = font_header
    ws.row_dimensions[1].height = 40

    keg_nama = df_header['Kegiatan'].iloc[0].title()
    keg_kode = kegiatan_code_map.get(df_header['Kegiatan'].iloc[0], "0000")

    meta_rows = [
        ("Kementerian/ Lembaga:", "(023) KEMENTERIAN PENDIDIKAN, KEBUDAYAAN, RISET DAN TEKNOLOGI"), 
        ("Unit Eselon II/ Satker:", "(17) Dirjen Diktiristek / (677567) UNIVERSITAS MULAWARMAN"),
        ("Sumber Dana:", df_header.get('Sumber_Dana', pd.Series(['BOPTN'])).iloc[0]),
        ("Kegiatan:", f"{keg_kode} - {keg_nama}"), ("Sasaran Kegiatan:", df_header['Sasaran'].iloc[0]), 
        ("Klasifikasi Rincian Output:", df_header['KRO'].iloc[0]), ("Volume:", df_header['Volume'].iloc[0]), 
        ("Satuan Ukur:", df_header['Satuan'].iloc[0]), ("Alokasi Dana:", f"Rp. {df_items['Total_Biaya'].sum():,.0f}".replace(',','.'))
    ]
    rp = 2
    for label, val in meta_rows:
        ws.cell(row=rp, column=1, value=label).font = font_bold
        ws.cell(row=rp, column=2, value=val)
        ws.merge_cells(start_row=rp, start_column=2, end_row=rp, end_column=5)
        rp += 1

    rp += 1
    for col_idx, text in enumerate(["Kode", "Rincian Belanja", "Volume & Satuan", "Harga Satuan", "Jumlah Biaya"], start=1):
        cell = ws.cell(row=rp, column=col_idx, value=text); cell.font = font_bold; cell.alignment = align_center; cell.border = border_all
    rp += 1

    def print_row(kode, urai, vol, hrg, tot, is_bold=False):
        nonlocal rp
        ws.cell(row=rp, column=1, value=kode).border = border_all
        ws.cell(row=rp, column=2, value=urai).border = border_all
        ws.cell(row=rp, column=3, value=vol).border = border_all
        ws.cell(row=rp, column=4, value=hrg).border = border_all
        if hrg != "": ws.cell(row=rp, column=4).number_format = '#,##0'
        ws.cell(row=rp, column=5, value=tot).border = border_all; ws.cell(row=rp, column=5).number_format = '#,##0'
        if is_bold: 
            for col in range(1,6): ws.cell(row=rp, column=col).font = Font(bold=True)
        rp += 1

    total_seluruh = df_items["Total_Biaya"].sum()
    for head_col, indent in [('RO', ""), ('Komponen', "  "), ('Sub_Komponen', "    ")]:
        if df_header[head_col].iloc[0] and str(df_header[head_col].iloc[0]).strip() not in ["", "-", "Tidak Ada Sub-Komponen"]:
            k_val, u_val = split_kode(df_header[head_col].iloc[0])
            print_row(k_val, f"{indent}{u_val}", "", "", total_seluruh, True)

    print_row(keg_kode, f"      {keg_nama}", "", "", total_seluruh, True)

    for akun, group in df_items.groupby("Akun_Belanja"):
        k_ak, u_ak = split_kode(akun)
        print_row(k_ak, f"        {u_ak}", "", "", group['Total_Biaya'].sum(), True)
        for _, r in group.iterrows():
            v_sat = get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2'])
            print_row("", f"          - {r['Uraian']}", v_sat, r['Harga_Satuan'], r['Total_Biaya'])
            
    rp += 2
    try: tgl_str = format_tgl_indo(df_header['Tgl_Cetak'].iloc[0])
    except: tgl_str = df_header['Tgl_Cetak'].iloc[0]
    
    ws.cell(row=rp, column=4, value=f"Samarinda, {tgl_str}")
    ws.cell(row=rp+1, column=4, value=df_header['Jabatan'].iloc[0])
    ws.cell(row=rp+5, column=4, value=df_header['Nama_Pejabat'].iloc[0]).font = Font(underline="single", bold=True)
    ws.cell(row=rp+6, column=4, value=f"NIP. {df_header['NIP_Pejabat'].iloc[0]}")

    if tampilkan_paraf:
        rp_paraf = rp
        ws.cell(row=rp_paraf, column=1, value="No").border = border_all; ws.cell(row=rp_paraf, column=1).font = font_bold; ws.cell(row=rp_paraf, column=1).alignment = align_center
        ws.cell(row=rp_paraf, column=2, value="Jabatan").border = border_all; ws.cell(row=rp_paraf, column=2).font = font_bold; ws.cell(row=rp_paraf, column=2).alignment = align_center
        ws.cell(row=rp_paraf, column=3, value="Paraf").border = border_all; ws.cell(row=rp_paraf, column=3).font = font_bold; ws.cell(row=rp_paraf, column=3).alignment = align_center
        jabatan_list = ["Wakil Dekan Bidang Keuangan dan Umum", "Kepala Bagian Umum", "Staf Perencanaan"]
        for i, jab in enumerate(jabatan_list, start=1):
            rp_paraf += 1
            ws.cell(row=rp_paraf, column=1, value=i).border = border_all; ws.cell(row=rp_paraf, column=1).alignment = align_center
            ws.cell(row=rp_paraf, column=2, value=f" {jab}").border = border_all
            ws.cell(row=rp_paraf, column=3, value="").border = border_all
            ws.row_dimensions[rp_paraf].height = 25

    output = BytesIO(); wb.save(output)
    return output.getvalue()

def export_pdf_rab(df_header, df_items, orientasi, kegiatan_code_map, tampilkan_paraf=False):
    total_seluruh = df_items["Total_Biaya"].sum()
    t_rab = df_header.get('Tahun', pd.Series(['2027'])).iloc[0]
    s_dana = df_header.get('Sumber_Dana', pd.Series(['BOPTN'])).iloc[0]
    keg_nama_full = df_header['Kegiatan'].iloc[0].title()
    keg_kode_full = kegiatan_code_map.get(df_header['Kegiatan'].iloc[0], "0000")
    
    try: tgl_str = f"Samarinda, {format_tgl_indo(df_header['Tgl_Cetak'].iloc[0])}"
    except: tgl_str = f"Samarinda, {df_header['Tgl_Cetak'].iloc[0]}"
    
    page_rule = "A4 landscape" if orientasi == "Landscape" else "A4 portrait"
    
    paraf_html = ""
    if tampilkan_paraf:
        paraf_html = """
        <table style="width: 320px; border-collapse: collapse; float: left; margin-top: 20px; font-size: 8pt;">
            <tr>
                <th style="border: 1px solid black; padding: 4px; text-align: center; width: 10%;">No</th>
                <th style="border: 1px solid black; padding: 4px; text-align: center; width: 65%;">Jabatan</th>
                <th style="border: 1px solid black; padding: 4px; text-align: center; width: 25%;">Paraf</th>
            </tr>
            <tr><td style="border: 1px solid black; height: 35px; text-align: center; vertical-align: middle;">1</td><td style="border: 1px solid black; padding-left: 5px;">Wakil Dekan Bidang Keuangan dan Umum</td><td style="border: 1px solid black;"></td></tr>
            <tr><td style="border: 1px solid black; height: 35px; text-align: center; vertical-align: middle;">2</td><td style="border: 1px solid black; padding-left: 5px;">Kepala Bagian Umum</td><td style="border: 1px solid black;"></td></tr>
            <tr><td style="border: 1px solid black; height: 35px; text-align: center; vertical-align: middle;">3</td><td style="border: 1px solid black; padding-left: 5px;">Staf Perencanaan</td><td style="border: 1px solid black;"></td></tr>
        </table>
        """

    html = f"""
    <!DOCTYPE html>
    <html><head><meta charset="utf-8">
    <style>
        @page {{ size: {page_rule}; margin: 10mm; }}
        body {{ font-family: 'Arial', sans-serif; font-size: 8.5pt; line-height: 1.2; color: #000; }}
        .judul {{ text-align: center; font-weight: bold; font-size: 11pt; margin-bottom: 15px; }}
        .tabel-meta td {{ padding: 1px 3px; font-size: 8.5pt; }}
        .tabel-utama {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 8pt; }}
        .tabel-utama th, .tabel-utama td {{ border: 1px solid black; padding: 4px; }}
        .tabel-utama th {{ background-color: #d9d9d9; text-align: center; font-weight: bold; }}
        .text-right {{ text-align: right; }} .text-center {{ text-align: center; }} .bold {{ font-weight: bold; }}
        .ttd-box {{ width: 220px; float: right; text-align: left; margin-top: 20px; margin-right: 15px; page-break-inside: avoid; }}
        .kro-row {{ background-color: #d9e1f2; }} .ro-row {{ background-color: #e9edf4; }}
        .komp-row {{ background-color: #fff2cc; }} .sub-row {{ background-color: #fce4d6; }}
        .keg-row {{ background-color: #e2efda; }} .akun-row {{ background-color: #ffffff; }}
    </style></head><body>
    <div class="judul">RINCIAN ANGGARAN BIAYA (RAB) FAKULTAS ILMU BUDAYA<br>TAHUN ANGGARAN {t_rab}</div>
    <table class="tabel-meta">
        <tr><td class="bold">Kementerian/ Lembaga</td><td>:</td><td>(023) KEMENTERIAN PENDIDIKAN, KEBUDAYAAN, RISET DAN TEKNOLOGI</td></tr>
        <tr><td class="bold">Unit Eselon II/ Satker</td><td>:</td><td>(17) Dirjen Diktiristek / (677567) UNIVERSITAS MULAWARMAN</td></tr>
        <tr><td class="bold">Sumber Dana</td><td>:</td><td>{s_dana}</td></tr>
        <tr><td class="bold">Kegiatan</td><td>:</td><td>{keg_kode_full} - {keg_nama_full}</td></tr>
        <tr><td class="bold">Sasaran Kegiatan</td><td>:</td><td>{df_header['Sasaran'].iloc[0]}</td></tr>
        <tr><td class="bold">Klasifikasi Rincian Output</td><td>:</td><td>{df_header['KRO'].iloc[0]}</td></tr>
        <tr><td class="bold">Volume</td><td>:</td><td>{df_header['Volume'].iloc[0]}</td></tr>
        <tr><td class="bold">Satuan Ukur</td><td>:</td><td>{df_header['Satuan'].iloc[0]}</td></tr>
        <tr><td class="bold">Alokasi Dana (Total Belanja)</td><td>:</td><td>Rp. {format_rupiah(total_seluruh)}</td></tr>
    </table>
    <table class="tabel-utama">
        <tr><th>Kode</th><th>Rincian Belanja</th><th>Volume & Satuan</th><th>Harga Satuan</th><th>Jumlah Biaya</th></tr>
    """
    for head_col, indent, cls_row in [('RO', "", "ro-row"), ('Komponen', "  ", "komp-row"), ('Sub_Komponen', "    ", "sub-row")]:
        if df_header[head_col].iloc[0] and str(df_header[head_col].iloc[0]).strip() not in ["", "-", "Tidak Ada Sub-Komponen"]:
            k, u = split_kode(df_header[head_col].iloc[0])
            html += f"<tr class='{cls_row} bold'><td>{k}</td><td>{indent}{u}</td><td></td><td></td><td class='right'>{format_rupiah(total_seluruh)}</td></tr>"
    
    html += f"<tr class='keg-row bold'><td>{keg_kode_full}</td><td style='padding-left:15px;'>{keg_nama_full}</td><td></td><td></td><td class='right'>{format_rupiah(total_seluruh)}</td></tr>"

    for akun, group_akun in df_items.groupby("Akun_Belanja"):
        k_ak, u_ak = split_kode(akun)
        html += f"<tr class='akun-row bold'><td>{k_ak}</td><td style='padding-left:30px;'>{u_ak}</td><td></td><td></td><td class='right'>{format_rupiah(group_akun['Total_Biaya'].sum())}</td></tr>"
        for _, r in group_akun.iterrows():
            v_sat_str = get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2'])
            html += f"<tr><td></td><td style='padding-left:45px;'>- {r['Uraian']}</td><td class='center'>{v_sat_str}</td><td class='right'>{format_rupiah(r['Harga_Satuan'])}</td><td class='right'>{format_rupiah(r['Total_Biaya'])}</td></tr>"
    
    html += f"""</table>
    <div class="ttd-box">
        {tgl_str}<br>{df_header['Jabatan'].iloc[0]}<br><br><br><br><br>
        <b><u>{df_header['Nama_Pejabat'].iloc[0]}</u></b><br>NIP. {df_header['NIP_Pejabat'].iloc[0]}
    </div>
    {paraf_html}
    <div style="clear: both;"></div>
    </body></html>"""
    return html

# --- UI LOGIC ---
if df_rab_utama.empty: 
    st.info(f"Belum ada dokumen RAB untuk Tahun {tahun_aktif}.")
else:
    col_a1, col_a2 = st.columns(2)
    versi_list_aktif = sorted(df_rab_utama['Versi_RAB'].unique())
    pilih_v_arsip = col_a1.selectbox("1. Pilih Versi (Untuk Difilter):", versi_list_aktif)
    sumber_dana_arsip = col_a2.radio("2. Pilih Sumber Dana:", ["BOPTN", "PNBP"], key="sd_arsip", horizontal=True)
    
    df_v_terpilih = df_rab_utama[(df_rab_utama['Versi_RAB'] == pilih_v_arsip) & (df_rab_utama['Sumber_Dana'] == sumber_dana_arsip)]
    
    active_versions = df_rab_utama[df_rab_utama['Is_Active'] == 1]['Versi_RAB'].unique()
    is_v_active = 1 if pilih_v_arsip in active_versions else 0
    
    if len(active_versions) > 1:
        st.error(f"🚨 **TERDETEKSI KONFLIK MULTI-VERSI:** Ada {len(active_versions)} versi yang berstatus Aktif secara bersamaan ({', '.join(active_versions)}).")
        if st.button(f"🛠️ Perbaiki & Jadikan HANYA Versi '{pilih_v_arsip}' Sebagai Acuan Utama", type="primary", use_container_width=True):
            df_rab_utama['Is_Active'] = 0
            df_rab_utama.loc[df_rab_utama['Versi_RAB'] == pilih_v_arsip, 'Is_Active'] = 1
            update_rab_tahun(df_rab_utama, df_rab_detail, tahun_aktif)
            log_audit("PERBAIKI VERSI", f"Memperbaiki konflik versi dan mengaktifkan versi '{pilih_v_arsip}'")
            st.rerun()
    elif is_v_active == 1:
        st.success(f"✅ **STATUS VERSI: AKTIF (FINAL ACUAN)**. Seluruh kegiatan pada versi '{pilih_v_arsip}' ditarik ke dalam Rekapitulasi Global RKAKL.")
    else:
        st.warning(f"🗄️ **STATUS VERSI: ARSIP REVISI (TIDAK AKTIF)**. Versi ini disimpan sebagai rekaman sejarah anggaran.")
        if st.button(f"🔄 Jadikan Seluruh Kegiatan di Versi '{pilih_v_arsip}' Sebagai Acuan Aktif", type="primary"):
            df_rab_utama['Is_Active'] = 0
            df_rab_utama.loc[df_rab_utama['Versi_RAB'] == pilih_v_arsip, 'Is_Active'] = 1
            update_rab_tahun(df_rab_utama, df_rab_detail, tahun_aktif)
            log_audit("AKTIFKAN VERSI", f"Menjadikan versi '{pilih_v_arsip}' sebagai versi aktif")
            st.toast(f"Versi {pilih_v_arsip} Berhasil Diaktifkan!"); st.rerun()

    st.markdown("---")
    
    with st.expander(f"⚠️ Zona Berbahaya: Hapus Seluruh Data Versi '{pilih_v_arsip}' ({sumber_dana_arsip})"):
        st.error("Tindakan ini akan menghapus PERMANEN seluruh kegiatan dan rincian pada versi dan sumber dana yang dipilih.")
        konfirmasi_hapus = st.text_input(f"Ketik 'HAPUS' (huruf besar) untuk melanjutkan penghapusan {pilih_v_arsip} - {sumber_dana_arsip}:")
        if st.button("🗑️ Eksekusi Hapus Versi", type="primary", use_container_width=True):
            if konfirmasi_hapus == "HAPUS":
                ids_to_delete = df_v_terpilih['ID_RAB'].tolist()
                df_rab_utama = df_rab_utama[~df_rab_utama['ID_RAB'].isin(ids_to_delete)]
                df_rab_detail = df_rab_detail[~df_rab_detail['ID_RAB'].isin(ids_to_delete)]
                update_rab_tahun(df_rab_utama, df_rab_detail, tahun_aktif)
                log_audit("HAPUS VERSI", f"Menghapus permanen seluruh data versi '{pilih_v_arsip}' ({sumber_dana_arsip})")
                st.success(f"Seluruh data versi {pilih_v_arsip} ({sumber_dana_arsip}) berhasil dihapus!"); st.rerun()
            else:
                st.warning("Konfirmasi gagal. Pastikan mengetik 'HAPUS' dengan huruf besar.")
    st.markdown("---")

    keg_list_aktif = sorted(df_v_terpilih['Kegiatan'].unique())
    pilih_keg_arsip = st.selectbox(f"3. Pilih Kegiatan (Dalam Versi {pilih_v_arsip} - {sumber_dana_arsip}):", keg_list_aktif, format_func=lambda x: x.title())
    
    with st.expander(f"📋 Duplikasi Seluruh Kegiatan Versi '{pilih_v_arsip}' ke Versi Lain"):
        target_versi = st.selectbox("Pilih Target Versi Baru:", ["Transisi","Indikatif", "Definitif", "Revisi 1", "Revisi 2", "Revisi 3", "Revisi 4", "Revisi 5", "Revisi 6", "Revisi 7", "Revisi 8", "Revisi 9", "Revisi 10","Revisi 11","Revisi 12","Revisi 13"])
        if st.button(f"🚀 Salin Semua ke '{target_versi}'", type="primary"):
            df_rab_utama['Is_Active'] = 0
            df_v_asal = df_rab_utama[df_rab_utama['Versi_RAB'] == pilih_v_arsip] 
            for i, (_, row_keg) in enumerate(df_v_asal.iterrows()):
                new_row = row_keg.copy()
                new_id = f"RAB-{datetime.now().strftime('%Y%m%d%H%M%S%f')}-{i}"
                new_row['ID_RAB'] = new_id
                new_row['Versi_RAB'] = target_versi
                new_row['Is_Active'] = 1
                df_rab_utama = pd.concat([df_rab_utama, pd.DataFrame([new_row])], ignore_index=True)
                
                old_id = row_keg['ID_RAB']
                det_keg_lama = df_rab_detail[df_rab_detail['ID_RAB'] == old_id].copy()
                det_keg_lama['ID_RAB'] = new_id
                df_rab_detail = pd.concat([df_rab_detail, det_keg_lama], ignore_index=True)
                
            update_rab_tahun(df_rab_utama, df_rab_detail, tahun_aktif)
            log_audit("DUPLIKASI VERSI", f"Menyalin seluruh kegiatan versi '{pilih_v_arsip}' menjadi versi '{target_versi}'")
            st.success(f"Berhasil menduplikasi ke {target_versi} dan versi tersebut langsung diaktifkan!"); st.rerun()

    if pilih_keg_arsip:
        head_terpilih = df_v_terpilih[df_v_terpilih['Kegiatan'] == pilih_keg_arsip]
        id_rab_aktif = head_terpilih['ID_RAB'].iloc[0]
        detail_terpilih = df_rab_detail[df_rab_detail["ID_RAB"] == id_rab_aktif]
        
        if st.button("✏️ Edit Kegiatan Ini", use_container_width=True, type="secondary"):
            st.session_state.edit_rab_id = id_rab_aktif
            st.switch_page("modul_rab/rab_buat.py")

        st.markdown("---")
        df_view = detail_terpilih.copy()
        df_view['Kode Akun'] = df_view['Akun_Belanja'].apply(lambda x: split_kode(x)[0])
        df_view['Nama Akun Belanja'] = df_view['Akun_Belanja'].apply(lambda x: split_kode(x)[1])
        df_view['Volume & Satuan'] = df_view.apply(lambda r: get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2']), axis=1)
        
        keg_code_view = kegiatan_code_map.get(pilih_keg_arsip, "0000")
        s_dana = head_terpilih.get('Sumber_Dana', pd.Series(['BOPTN'])).iloc[0]
        st.markdown(f"**Identitas Kegiatan:** {keg_code_view} - {pilih_keg_arsip.title()}")
        st.markdown(f"**Klasifikasi Dokumen:** {head_terpilih['KRO'].iloc[0]} ➔ {head_terpilih['RO'].iloc[0]}")
        st.markdown(f"**Total Alokasi Anggaran ({s_dana}):** Rp {format_rupiah(detail_terpilih['Total_Biaya'].sum())}")
        st.markdown(f"**Catatan Revisi:** {head_terpilih.get('Catatan', pd.Series(['-'])).iloc[0]}")
        st.dataframe(df_view[["Kode Akun", "Nama Akun Belanja", "Uraian", "Volume & Satuan", "Harga_Satuan", "Total_Biaya"]].style.format({"Harga_Satuan": format_rupiah, "Total_Biaya": format_rupiah}), hide_index=True, use_container_width=True)

        st.markdown("#### 🖨️ Cetak Dokumen Satuan")
        tampilkan_paraf_rab = st.checkbox("Tampilkan Tabel Paraf (Khusus Arsip Hardcopy Internal)", key=f"paraf_{id_rab_aktif}")
        
        col_x1, col_x2, col_x3 = st.columns([1, 1, 2])
        with col_x1:
            st.download_button("📥 Download Excel Resmi", data=export_excel_rab(head_terpilih, detail_terpilih, kegiatan_code_map, tampilkan_paraf_rab), file_name=f"RAB_{s_dana}_{pilih_keg_arsip}_{pilih_v_arsip}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col_x2:
            st.download_button("📑 Download PDF (Web)", data=export_pdf_rab(head_terpilih, detail_terpilih, "Portrait", kegiatan_code_map, tampilkan_paraf_rab).encode('utf-8'), file_name=f"RAB_{s_dana}_{pilih_keg_arsip}_{pilih_v_arsip}.html", mime="text/html", use_container_width=True)
        with col_x3:
            if st.button("🗑️ Hapus Dokumen Ini", type="secondary", use_container_width=True):
                df_rab_utama = df_rab_utama[df_rab_utama["ID_RAB"] != id_rab_aktif]
                df_rab_detail = df_rab_detail[df_rab_detail["ID_RAB"] != id_rab_aktif]
                update_rab_tahun(df_rab_utama, df_rab_detail, tahun_aktif)
                log_audit("HAPUS DOKUMEN", f"Menghapus dokumen kegiatan tunggal: '{pilih_keg_arsip}'")
                st.rerun()
