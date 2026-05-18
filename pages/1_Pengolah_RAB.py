import streamlit as st
import pandas as pd
from io import BytesIO

# Konfigurasi Halaman (Harus selalu ada di baris pertama)
st.set_page_config(page_title="Pengolah Dokumen RAB", page_icon="📄", layout="wide")

# ==========================================
# 🛡️ PENGAMAN HALAMAN (SECURITY LOCK)
# ==========================================
# Cek apakah user sudah login
if "logged_in" not in st.session_state or not st.session_state["logged_in"]:
    st.warning("🔒 Sesi Anda telah berakhir atau Anda belum Login. Silakan kembali ke Halaman Utama untuk Login.")
    st.stop() # Hentikan kode agar tidak menampilkan apapun

# Cek apakah user adalah ADMIN (Hanya admin yang boleh buka halaman ini)
if st.session_state.get("role") != "admin":
    st.error("🚫 Akses Ditolak! Halaman ini dikhususkan untuk Administrator Fakultas Ilmu Budaya.")
    st.stop() # Hentikan eksekusi kode

# ==========================================
# SIDEBAR
# ==========================================
with st.sidebar:
    st.header("Sistem Perencanaan")
    st.markdown(f"👤 **{st.session_state['nama_user']}**")
    
    st.markdown("---")
    st.info("💡 Anda sedang berada di halaman **Pengolah Dokumen RAB**. Untuk kembali ke laporan Prodi, klik menu **'kompiler prodi'** di atas.")
    
    if st.button("🚪 Logout", type="primary"):
        st.session_state.update({"logged_in": False, "role": None, "nama_user": None, "username": None})
        st.switch_page("kompiler_prodi.py") # Tendang kembali ke halaman utama saat logout

# ==========================================
# APLIKASI PENGOLAH RAB (HANYA MUNCUL JIKA ADMIN)
# ==========================================
st.title("📄 Pengolah Dokumen RAB (Otomatis)")
st.info("Aplikasi khusus untuk memproduksi file Excel resmi (RAB) sesuai standar format universitas. Isi data form di bawah, lalu klik 'Buat File RAB'.")

# SECTION 1: HEADER RAB
st.markdown("### 1. Informasi Utama RAB")
col1, col2 = st.columns(2)

st.text_input("Kementerian/ Lembaga (Otomatis)", "(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI", disabled=True)
st.text_input("Unit Eselon II/ Satker (Otomatis)", "(17) Direktorat Jenderal Pendidikan Tinggi / (677524) UNIVERSITAS MULAWARMAN", disabled=True)

rab_kegiatan = col1.text_input("Kegiatan", "(7730) PENGELOLAAN BOPTN")
rab_sasaran = col2.text_input("Sasaran Kegiatan", "Meningkatnya Kualitas Lulusan Pendidikan Tinggi")
rab_kro = col1.text_input("Klasifikasi Rincian Output", "(7730.CAA) Layanan Dukungan Manajemen Internal")
rab_vol = col2.number_input("Volume", value=1, min_value=1)
rab_satuan = col1.text_input("Satuan Ukur", "Layanan")
rab_alokasi = col2.number_input("Total Alokasi Dana (Rp)", value=12000000, step=1000000)

st.markdown("---")
# SECTION 2: DETAIL BELANJA
st.markdown("### 2. Rincian Anggaran (Tabel Belanja)")
st.caption("Masukkan rincian item. Mesin akan otomatis mengelompokkannya ke dalam format Excel yang berjenjang.")

template_rab = pd.DataFrame([{
    "Rincian Output": "7730.CAA.001 - Layanan Perkantoran",
    "Komponen": "052.A - Penyelenggaraan Rapat Koordinasi Kurikulum",
    "Akun Belanja": "525114 - Belanja Barang Persediaan",
    "Uraian Belanja (Barang/Jasa)": "Konsumsi Rapat",
    "Volume": 30,
    "Satuan": "Orang",
    "Harga Satuan": 400000
}])

df_rab_input = st.data_editor(
    template_rab,
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True,
    column_config={
        "Rincian Output": st.column_config.SelectboxColumn("Pilih Rincian Output", options=["7730.CAA.001 - Layanan Perkantoran", "7730.CAA.002 - Pemeliharaan Gedung"], required=True),
        "Komponen": st.column_config.SelectboxColumn("Pilih Komponen", options=["052.A - Penyelenggaraan Rapat Koordinasi Kurikulum", "051.A - Pemeliharaan Rutin"], required=True),
        "Akun Belanja": st.column_config.SelectboxColumn("Pilih Akun Belanja", options=["525114 - Belanja Barang Persediaan", "521111 - Belanja Keperluan Perkantoran"], required=True),
        "Uraian Belanja (Barang/Jasa)": st.column_config.TextColumn("Detail/Uraian", required=True),
        "Volume": st.column_config.NumberColumn("Volume", min_value=1, required=True),
        "Satuan": st.column_config.SelectboxColumn("Satuan", options=["Orang", "Paket", "Kegiatan", "Bulan", "Lembar", "Box"], required=True),
        "Harga Satuan": st.column_config.NumberColumn("Harga Satuan (Rp)", min_value=0, required=True)
    }
)

def export_excel_rab(df_items, header_data):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB 2027"

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    font_bold = Font(bold=True)
    font_header = Font(bold=True, size=12)
    align_center = Alignment(horizontal="center", vertical="center")
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws["A1"] = "RINCIAN ANGGARAN BIAYA (RAB) TAHUN 2027"
    ws["A1"].font = font_header

    meta_rows = [
        ("Kementerian/ Lembaga:", "(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI"),
        ("Unit Eselon II/ Satker:", "(17) Direktorat Jenderal Pendidikan Tinggi / (677524) UNIVERSITAS MULAWARMAN"),
        ("Kegiatan:", header_data['kegiatan']),
        ("Sasaran Kegiatan:", header_data['sasaran']),
        ("Klasifikasi Rincian Output:", header_data['kro']),
        ("Volume:", header_data['vol']),
        ("Satuan Ukur:", header_data['satuan']),
        ("Alokasi Dana:", f"Rp. {header_data['alokasi']:,.0f}".replace(',','.'))
    ]

    row_pointer = 2
    for label, val in meta_rows:
        ws.cell(row=row_pointer, column=1, value=label).font = font_bold
        ws.cell(row=row_pointer, column=2, value=val)
        row_pointer += 1

    row_pointer += 1
    headers_tabel = ["Rincian", "Volume", "Satuan", "Harga Satuan", "Jumlah Biaya"]
    for col_idx, text in enumerate(headers_tabel, start=1):
        cell = ws.cell(row=row_pointer, column=col_idx, value=text)
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_all

    row_pointer += 1
    df_items['Jumlah_Biaya'] = df_items['Volume'] * df_items['Harga Satuan']

    for ro, group_ro in df_items.groupby("Rincian Output"):
        tot_ro = group_ro["Jumlah_Biaya"].sum()
        cell_ro1 = ws.cell(row=row_pointer, column=1, value=ro)
        cell_ro1.font = font_bold; cell_ro1.border = border_all
        ws.cell(row=row_pointer, column=2).border = border_all
        ws.cell(row=row_pointer, column=3).border = border_all
        ws.cell(row=row_pointer, column=4).border = border_all
        cell_ro5 = ws.cell(row=row_pointer, column=5, value=tot_ro)
        cell_ro5.font = font_bold; cell_ro5.border = border_all; cell_ro5.number_format = '#,##0'
        row_pointer += 1

        for komp, group_komp in group_ro.groupby("Komponen"):
            tot_komp = group_komp["Jumlah_Biaya"].sum()
            cell_ko1 = ws.cell(row=row_pointer, column=1, value=f"  {komp}")
            cell_ko1.font = font_bold; cell_ko1.border = border_all
            ws.cell(row=row_pointer, column=2).border = border_all
            ws.cell(row=row_pointer, column=3).border = border_all
            ws.cell(row=row_pointer, column=4).border = border_all
            cell_ko5 = ws.cell(row=row_pointer, column=5, value=tot_komp)
            cell_ko5.font = font_bold; cell_ko5.border = border_all; cell_ko5.number_format = '#,##0'
            row_pointer += 1

            for akun, group_akun in group_komp.groupby("Akun Belanja"):
                tot_akun = group_akun["Jumlah_Biaya"].sum()
                cell_ak1 = ws.cell(row=row_pointer, column=1, value=f"    {akun}")
                cell_ak1.font = font_bold; cell_ak1.border = border_all
                ws.cell(row=row_pointer, column=2).border = border_all
                ws.cell(row=row_pointer, column=3).border = border_all
                ws.cell(row=row_pointer, column=4).border = border_all
                cell_ak5 = ws.cell(row=row_pointer, column=5, value=tot_akun)
                cell_ak5.font = font_bold; cell_ak5.border = border_all; cell_ak5.number_format = '#,##0'
                row_pointer += 1

                for _, baris_data in group_akun.iterrows():
                    c_ur = ws.cell(row=row_pointer, column=1, value=f"      - {baris_data['Uraian Belanja (Barang/Jasa)']}")
                    c_ur.border = border_all
                    c_vol = ws.cell(row=row_pointer, column=2, value=baris_data['Volume']); c_vol.alignment = align_center; c_vol.border = border_all
                    c_sat = ws.cell(row=row_pointer, column=3, value=baris_data['Satuan']); c_sat.alignment = align_center; c_sat.border = border_all
                    c_hrg = ws.cell(row=row_pointer, column=4, value=baris_data['Harga Satuan']); c_hrg.number_format = '#,##0'; c_hrg.border = border_all
                    c_tot = ws.cell(row=row_pointer, column=5, value=baris_data['Jumlah_Biaya']); c_tot.number_format = '#,##0'; c_tot.border = border_all
                    row_pointer += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

if st.button("🚀 Buat File RAB (Excel)", type="primary"):
    valid_rab = df_rab_input[df_rab_input["Uraian Belanja (Barang/Jasa)"].str.strip() != ""]
    if valid_rab.empty:
        st.error("Gagal! Anda belum memasukkan rincian belanja sama sekali.")
    else:
        header_metadata = {
            "kegiatan": rab_kegiatan, "sasaran": rab_sasaran, "kro": rab_kro,
            "vol": rab_vol, "satuan": rab_satuan, "alokasi": rab_alokasi
        }
        file_excel_final = export_excel_rab(valid_rab, header_metadata)
        
        st.success("✅ File RAB Berhasil Dibuat dan Disusun Berjenjang!")
        st.download_button(
            label="📥 Download RAB Universitas.xlsx",
            data=file_excel_final,
            file_name="RAB_FIB_Resmi_2027.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
