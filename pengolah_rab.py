import streamlit as st
import pandas as pd
import os
import sqlite3
from io import BytesIO
from datetime import datetime

# ==========================================
# 🛡️ PENGAMAN HALAMAN (SECURITY LOCK)
# ==========================================
if "logged_in" not in st.session_state or not st.session_state["logged_in"]:
    st.warning("🔒 Sesi Anda telah berakhir atau Anda belum Login. Silakan kembali ke Halaman Utama untuk Login.")
    st.stop()

if st.session_state.get("role") != "admin":
    st.error("🚫 Akses Ditolak! Halaman ini dikhususkan untuk Administrator Fakultas Ilmu Budaya.")
    st.stop()

# ==========================================
# 🗄️ KONFIGURASI DATABASE RAB
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_DB = os.path.join(BASE_DIR, "database_usulan_prodi.db")

def load_table(table_name, default_cols):
    conn = sqlite3.connect(FILE_DB)
    try:
        df = pd.read_sql(f"SELECT * FROM {table_name}", conn)
    except:
        df = pd.DataFrame(columns=default_cols)
        df.to_sql(table_name, conn, index=False)
    conn.close()
    return df

def save_table(df, table_name):
    conn = sqlite3.connect(FILE_DB)
    df.to_sql(table_name, conn, if_exists="replace", index=False)
    conn.close()

# Muat Data Master
df_m_kro = load_table("rab_m_kro", ["KRO"])
df_m_ro = load_table("rab_m_ro", ["KRO", "RO"])
df_m_komp = load_table("rab_m_komp", ["RO", "Komponen"])
df_m_akun = load_table("rab_m_akun", ["Akun_Belanja"])

# Muat Data Arsip RAB
df_rab_utama = load_table("rab_utama", ["ID_RAB", "Tanggal", "KRO", "RO", "Komponen", "Kegiatan", "Sasaran", "Volume", "Satuan", "Alokasi"])
df_rab_detail = load_table("rab_detail", ["ID_RAB", "Akun_Belanja", "Uraian", "Volume", "Satuan", "Harga_Satuan", "Total_Biaya"])

def format_rupiah(x):
    try:
        return f"{float(x):,.0f}".replace(',', '.')
    except (ValueError, TypeError):
        return x

# ==========================================
# SIDEBAR
# ==========================================
with st.sidebar:
    st.info("💡 Anda sedang berada di halaman **Pengolah Dokumen RAB**. Untuk kembali ke laporan Prodi, klik menu navigasi di atas.")
    if st.button("🚪 Logout dari RAB", type="primary"):
        st.session_state.update({"logged_in": False, "role": None, "nama_user": None, "username": None})
        st.rerun()

# ==========================================
# TAMPILAN UTAMA APLIKASI
# ==========================================
st.title("📄 Pengolah Dokumen RAB Universitas")
st.caption("Sistem Manajemen & Generator Rincian Anggaran Biaya (RAB) Berjenjang.")

tab_master, tab_buat, tab_daftar = st.tabs(["🗂️ Master Database", "📝 Buat RAB Baru", "📂 Daftar RAB Tersimpan"])

# ---------------------------------------------------------
# TAB 1: MASTER DATABASE (HIERARKI)
# ---------------------------------------------------------
with tab_master:
    st.subheader("Pengaturan Master Data RAB")
    st.info("💡 Input data di sini terlebih dahulu. Hierarkinya adalah: **KRO ➔ RO ➔ Komponen**. Akun Belanja berdiri sendiri.")
    
    col_m1, col_m2 = st.columns(2)
    
    with col_m1:
        st.markdown("**1. Master KRO (Klasifikasi Rincian Output)**")
        edit_kro = st.data_editor(df_m_kro, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_kro")
        if st.button("💾 Simpan KRO"):
            save_table(edit_kro.dropna(how='all'), "rab_m_kro")
            st.success("Master KRO Disimpan!"); st.rerun()
            
        st.markdown("**3. Master Komponen**")
        list_ro = df_m_ro["RO"].tolist() if not df_m_ro.empty else ["Isi Master RO Dulu"]
        edit_komp = st.data_editor(df_m_komp, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_komp",
                                  column_config={"RO": st.column_config.SelectboxColumn("Induk RO", options=list_ro, required=True)})
        if st.button("💾 Simpan Komponen"):
            save_table(edit_komp.dropna(how='all'), "rab_m_komp")
            st.success("Master Komponen Disimpan!"); st.rerun()

    with col_m2:
        st.markdown("**2. Master RO (Rincian Output)**")
        list_kro = df_m_kro["KRO"].tolist() if not df_m_kro.empty else ["Isi Master KRO Dulu"]
        edit_ro = st.data_editor(df_m_ro, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_ro",
                                 column_config={"KRO": st.column_config.SelectboxColumn("Induk KRO", options=list_kro, required=True)})
        if st.button("💾 Simpan RO"):
            save_table(edit_ro.dropna(how='all'), "rab_m_ro")
            st.success("Master RO Disimpan!"); st.rerun()
            
        st.markdown("**4. Master Akun Belanja**")
        edit_akun = st.data_editor(df_m_akun, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_akun")
        if st.button("💾 Simpan Akun Belanja"):
            save_table(edit_akun.dropna(how='all'), "rab_m_akun")
            st.success("Master Akun Disimpan!"); st.rerun()


# ---------------------------------------------------------
# TAB 2: BUAT RAB BARU
# ---------------------------------------------------------
with tab_buat:
    if df_m_kro.empty or df_m_ro.empty or df_m_komp.empty or df_m_akun.empty:
        st.warning("⚠️ Master Database masih kosong! Silakan isi data KRO, RO, Komponen, dan Akun Belanja di tab **Master Database** terlebih dahulu.")
    else:
        st.subheader("1. Klasifikasi Output RAB")
        col_c1, col_c2, col_c3 = st.columns(3)
        
        # Logika Dropdown Berjenjang
        pilih_kro = col_c1.selectbox("Pilih Klasifikasi Rincian Output (KRO)", df_m_kro["KRO"].tolist())
        
        # Filter RO berdasarkan KRO yang dipilih
        opsi_ro = df_m_ro[df_m_ro["KRO"] == pilih_kro]["RO"].tolist()
        pilih_ro = col_c2.selectbox("Pilih Rincian Output (RO)", opsi_ro if opsi_ro else ["Tidak ada RO untuk KRO ini"])
        
        # Filter Komponen berdasarkan RO yang dipilih
        opsi_komp = df_m_komp[df_m_komp["RO"] == pilih_ro]["Komponen"].tolist()
        pilih_komp = col_c3.selectbox("Pilih Komponen", opsi_komp if opsi_komp else ["Tidak ada Komponen untuk RO ini"])

        st.markdown("---")
        st.subheader("2. Informasi Utama Kegiatan")
        col_u1, col_u2 = st.columns(2)
        rab_kegiatan = col_u1.text_input("Nama Kegiatan", placeholder="Contoh: (7730) PENGELOLAAN BOPTN")
        rab_sasaran = col_u2.text_input("Sasaran Kegiatan", placeholder="Contoh: Meningkatnya Kualitas Lulusan...")
        rab_vol = col_u1.number_input("Volume Target", value=1, min_value=1)
        rab_satuan = col_u2.text_input("Satuan Ukur", placeholder="Contoh: Layanan / Dokumen")
        rab_alokasi = st.number_input("Total Pagu / Alokasi Dana (Rp)", value=0, step=1000000)

        st.markdown("---")
        st.subheader("3. Rincian Belanja")
        st.caption("Masukkan detail barang/jasa. Kolom 'Akun Belanja' diambil otomatis dari Master Database.")
        
        opsi_akun = df_m_akun["Akun_Belanja"].tolist()
        template_detail = pd.DataFrame([{"Akun Belanja": opsi_akun[0] if opsi_akun else "", "Uraian Belanja": "", "Volume": 1, "Satuan": "Paket", "Harga Satuan": 0}])
        
        df_input_detail = st.data_editor(
            template_detail, num_rows="dynamic", use_container_width=True, hide_index=True,
            column_config={
                "Akun Belanja": st.column_config.SelectboxColumn("Akun Belanja", options=opsi_akun, required=True),
                "Uraian Belanja": st.column_config.TextColumn("Uraian / Detail Barang", required=True),
                "Volume": st.column_config.NumberColumn("Volume", min_value=1, required=True),
                "Satuan": st.column_config.TextColumn("Satuan", required=True),
                "Harga Satuan": st.column_config.NumberColumn("Harga Satuan (Rp)", min_value=0, required=True)
            }
        )
        
        if st.button("💾 Simpan & Terbitkan RAB", type="primary"):
            valid_detail = df_input_detail[df_input_detail["Uraian Belanja"].str.strip() != ""]
            
            if not rab_kegiatan or valid_detail.empty:
                st.error("Gagal! Pastikan Nama Kegiatan dan minimal 1 Uraian Belanja terisi.")
            else:
                # Buat ID Unik untuk RAB ini
                id_rab_baru = f"RAB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                tgl_sekarang = datetime.now().strftime('%Y-%m-%d %H:%M')
                
                # Simpan Header ke database
                new_utama = pd.DataFrame([{
                    "ID_RAB": id_rab_baru, "Tanggal": tgl_sekarang, "KRO": pilih_kro, "RO": pilih_ro, "Komponen": pilih_komp,
                    "Kegiatan": rab_kegiatan, "Sasaran": rab_sasaran, "Volume": rab_vol, "Satuan": rab_satuan, "Alokasi": rab_alokasi
                }])
                df_rab_utama = pd.concat([df_rab_utama, new_utama], ignore_index=True)
                save_table(df_rab_utama, "rab_utama")
                
                # Simpan Detail ke database
                valid_detail["ID_RAB"] = id_rab_baru
                valid_detail["Total_Biaya"] = valid_detail["Volume"] * valid_detail["Harga Satuan"]
                valid_detail.rename(columns={"Akun Belanja": "Akun_Belanja", "Uraian Belanja": "Uraian", "Harga Satuan": "Harga_Satuan"}, inplace=True)
                
                df_rab_detail = pd.concat([df_rab_detail, valid_detail[["ID_RAB", "Akun_Belanja", "Uraian", "Volume", "Satuan", "Harga_Satuan", "Total_Biaya"]]], ignore_index=True)
                save_table(df_rab_detail, "rab_detail")
                
                st.success(f"✅ RAB '{rab_kegiatan}' Berhasil Disimpan!")
                st.rerun()


# ---------------------------------------------------------
# TAB 3: DAFTAR RAB TERSIMPAN (VIEW & EXCEL EXPORT)
# ---------------------------------------------------------
def export_excel_rab(df_header, df_items):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB Export"

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    font_bold = Font(bold=True)
    font_header = Font(bold=True, size=12)
    align_center = Alignment(horizontal="center", vertical="center")
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws["A1"] = "RINCIAN ANGGARAN BIAYA (RAB) TAHUN 2027"
    ws["A1"].font = font_header

    meta_rows = [
        ("Kementerian/ Lembaga:", "(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI"),
        ("Unit Eselon II/ Satker:", "(17) Direktorat Jenderal Pendidikan Tinggi / (677524) UNIVERSITAS MULAWARMAN"),
        ("Kegiatan:", df_header['Kegiatan'].iloc[0]),
        ("Sasaran Kegiatan:", df_header['Sasaran'].iloc[0]),
        ("Klasifikasi Rincian Output:", df_header['KRO'].iloc[0]),
        ("Volume:", df_header['Volume'].iloc[0]),
        ("Satuan Ukur:", df_header['Satuan'].iloc[0]),
        ("Alokasi Dana:", f"Rp. {df_header['Alokasi'].iloc[0]:,.0f}".replace(',','.'))
    ]

    row_pointer = 2
    for label, val in meta_rows:
        ws.cell(row=row_pointer, column=1, value=label).font = font_bold
        ws.cell(row=row_pointer, column=2, value=val)
        row_pointer += 1

    row_pointer += 1
    headers_tabel = ["Rincian", "Volume", "Satuan", "Harga Satuan", "Jumlah Biaya"]
    for col_idx, text in enumerate(headers_tabel, start=1):
        cell = ws.cell(row=row_pointer, column=col_idx, value=text)
        cell.font = font_bold; cell.alignment = align_center; cell.border = border_all

    row_pointer += 1
    
    # KARENA RAB INI SPESIFIK 1 RO DAN 1 KOMPONEN DARI HEADER
    ro_text = df_header['RO'].iloc[0]
    komp_text = df_header['Komponen'].iloc[0]
    total_seluruh = df_items["Total_Biaya"].sum()
    
    # Print baris RO
    c_ro1 = ws.cell(row=row_pointer, column=1, value=ro_text); c_ro1.font = font_bold; c_ro1.border = border_all
    ws.cell(row=row_pointer, column=2).border = border_all; ws.cell(row=row_pointer, column=3).border = border_all; ws.cell(row=row_pointer, column=4).border = border_all
    c_rot = ws.cell(row=row_pointer, column=5, value=total_seluruh); c_rot.font = font_bold; c_rot.border = border_all; c_rot.number_format = '#,##0'
    row_pointer += 1

    # Print baris Komponen
    c_ko1 = ws.cell(row=row_pointer, column=1, value=f"  {komp_text}"); c_ko1.font = font_bold; c_ko1.border = border_all
    ws.cell(row=row_pointer, column=2).border = border_all; ws.cell(row=row_pointer, column=3).border = border_all; ws.cell(row=row_pointer, column=4).border = border_all
    c_kot = ws.cell(row=row_pointer, column=5, value=total_seluruh); c_kot.font = font_bold; c_kot.border = border_all; c_kot.number_format = '#,##0'
    row_pointer += 1

    # Print Akun Belanja Grouping
    for akun, group_akun in df_items.groupby("Akun_Belanja"):
        tot_akun = group_akun["Total_Biaya"].sum()
        c_ak1 = ws.cell(row=row_pointer, column=1, value=f"    {akun}"); c_ak1.font = font_bold; c_ak1.border = border_all
        ws.cell(row=row_pointer, column=2).border = border_all; ws.cell(row=row_pointer, column=3).border = border_all; ws.cell(row=row_pointer, column=4).border = border_all
        c_akt = ws.cell(row=row_pointer, column=5, value=tot_akun); c_akt.font = font_bold; c_akt.border = border_all; c_akt.number_format = '#,##0'
        row_pointer += 1

        for _, baris_data in group_akun.iterrows():
            ws.cell(row=row_pointer, column=1, value=f"      - {baris_data['Uraian']}").border = border_all
            c_vol = ws.cell(row=row_pointer, column=2, value=baris_data['Volume']); c_vol.alignment = align_center; c_vol.border = border_all
            c_sat = ws.cell(row=row_pointer, column=3, value=baris_data['Satuan']); c_sat.alignment = align_center; c_sat.border = border_all
            c_hrg = ws.cell(row=row_pointer, column=4, value=baris_data['Harga_Satuan']); c_hrg.number_format = '#,##0'; c_hrg.border = border_all
            c_tot = ws.cell(row=row_pointer, column=5, value=baris_data['Total_Biaya']); c_tot.number_format = '#,##0'; c_tot.border = border_all
            row_pointer += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


with tab_daftar:
    if df_rab_utama.empty:
        st.info("Belum ada dokumen RAB yang tersimpan.")
    else:
        st.subheader("Arsip Dokumen RAB")
        
        # Buat label rapi untuk dropdown pilihan RAB
        opsi_arsip = {row['ID_RAB']: f"[{row['Tanggal']}] {row['Kegiatan']}" for _, row in df_rab_utama.iterrows()}
        pilih_arsip = st.selectbox("Pilih RAB yang ingin dilihat/diunduh:", options=list(opsi_arsip.keys()), format_func=lambda x: opsi_arsip[x])
        
        # Tampilkan Detail
        head_terpilih = df_rab_utama[df_rab_utama["ID_RAB"] == pilih_arsip]
        detail_terpilih = df_rab_detail[df_rab_detail["ID_RAB"] == pilih_arsip]
        
        st.markdown(f"**Klasifikasi:** {head_terpilih['KRO'].iloc[0]} ➔ {head_terpilih['RO'].iloc[0]} ➔ {head_terpilih['Komponen'].iloc[0]}")
        st.markdown(f"**Total Anggaran Disusun:** Rp {format_rupiah(detail_terpilih['Total_Biaya'].sum())} _(Pagu: Rp {format_rupiah(head_terpilih['Alokasi'].iloc[0])})_")
        
        st.dataframe(detail_terpilih[["Akun_Belanja", "Uraian", "Volume", "Satuan", "Harga_Satuan", "Total_Biaya"]].style.format({
            "Harga_Satuan": format_rupiah, "Total_Biaya": format_rupiah
        }), hide_index=True, use_container_width=True)
        
        col_x1, col_x2 = st.columns([1, 4])
        with col_x1:
            file_xls = export_excel_rab(head_terpilih, detail_terpilih)
            st.download_button("📥 Download Excel Resmi", data=file_xls, file_name=f"RAB_{pilih_arsip}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
        with col_x2:
            if st.button("🗑️ Hapus Dokumen Ini"):
                df_rab_utama = df_rab_utama[df_rab_utama["ID_RAB"] != pilih_arsip]
                df_rab_detail = df_rab_detail[df_rab_detail["ID_RAB"] != pilih_arsip]
                save_table(df_rab_utama, "rab_utama")
                save_table(df_rab_detail, "rab_detail")
                st.success("Dokumen terhapus!"); st.rerun()
