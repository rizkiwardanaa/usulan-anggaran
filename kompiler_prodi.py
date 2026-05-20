import streamlit as st
import pandas as pd
import os
import sqlite3
from io import BytesIO
from datetime import datetime

# ==========================================
# 1. KONFIGURASI HALAMAN & LOKASI DATABASE
# ==========================================
st.set_page_config(page_title="Kompiler Usulan Anggaran FIB", page_icon="📝", layout="wide")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_DB = os.path.join(BASE_DIR, "database_usulan_prodi.db")
FILE_CSV_LAMA = os.path.join(BASE_DIR, "database_usulan_prodi.csv")
UPLOAD_DIR = os.path.join(BASE_DIR, "tor_uploads")

if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# --- FUNGSI DATABASE KOMPILER PRODI ---
def load_data():
    conn = sqlite3.connect(FILE_DB)
    cek_tabel = pd.read_sql("SELECT name FROM sqlite_master WHERE type='table' AND name='usulan'", conn)
    
    if cek_tabel.empty:
        if os.path.exists(FILE_CSV_LAMA):
            df = pd.read_csv(FILE_CSV_LAMA)
            if "Status" not in df.columns: df["Status"] = "Menunggu Review"
            if "Catatan_Fakultas" not in df.columns: df["Catatan_Fakultas"] = "-"
            if "File_TOR" not in df.columns: df["File_TOR"] = "-"
            df.to_sql("usulan", conn, if_exists="replace", index=False)
            conn.close()
            return df
        else:
            df_kosong = pd.DataFrame(columns=["Tanggal_Input", "Program_Studi", "Nama_Kegiatan", "Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan", "Total_Usulan", "Prioritas", "Status", "Catatan_Fakultas", "File_TOR"])
            df_kosong.to_sql("usulan", conn, if_exists="replace", index=False)
            conn.close()
            return df_kosong
    else:
        df = pd.read_sql("SELECT * FROM usulan", conn)
        conn.close()
        return df

def save_data(df):
    conn = sqlite3.connect(FILE_DB)
    df.to_sql("usulan", conn, if_exists="replace", index=False)
    conn.close()

df_usulan = load_data()

# --- FUNGSI DATABASE MASTER RAB ---
def load_table(table_name, default_cols):
    conn = sqlite3.connect(FILE_DB)
    try:
        df = pd.read_sql(f"SELECT * FROM {table_name}", conn)
        for col in default_cols:
            if col not in df.columns:
                if "Vol" in col or "Harga" in col or "Total" in col: df[col] = 1 if "Vol" in col else 0
                elif col == "Tahun": df[col] = "2027"
                else: df[col] = "-"
    except:
        df = pd.DataFrame(columns=default_cols)
        df.to_sql(table_name, conn, index=False)
    conn.close()
    return df

def save_table(df, table_name):
    conn = sqlite3.connect(FILE_DB)
    df.to_sql(table_name, conn, if_exists="replace", index=False)
    conn.close()

df_m_kro = load_table("rab_m_kro", ["KRO"])
df_m_ro = load_table("rab_m_ro", ["KRO", "RO"])
df_m_komp = load_table("rab_m_komp", ["RO", "Komponen"])
df_m_subkomp = load_table("rab_m_subkomp", ["Komponen", "Sub_Komponen"])
df_m_akun = load_table("rab_m_akun", ["Account_Code", "Account_Name"]) 
df_m_pejabat = load_table("rab_m_pejabat", ["Jabatan", "Nama", "NIP"])

df_rab_utama = load_table("rab_utama", ["ID_RAB", "Tanggal", "Tahun", "Tgl_Cetak", "KRO", "RO", "Komponen", "Sub_Komponen", "Kegiatan", "Sasaran", "Volume", "Satuan", "Alokasi", "Jabatan", "Nama_Pejabat", "NIP_Pejabat"])
df_rab_detail = load_table("rab_detail", ["ID_RAB", "Akun_Belanja", "Uraian", "Vol_1", "Sat_1", "Vol_2", "Sat_2", "Harga_Satuan", "Total_Biaya"])

def format_rupiah(x):
    try: return f"{float(x):,.0f}".replace(',', '.')
    except (ValueError, TypeError): return x

def split_kode(teks):
    s = str(teks).strip()
    if " - " in s:
        parts = s.split(" - ", 1)
        return parts[0].strip(), parts[1].strip()
    parts = s.split(" ", 1)
    if len(parts) == 2:
        first_part = parts[0].strip()
        if any(c.isdigit() for c in first_part) or len(first_part) <= 8 or "." in first_part:
            return first_part, parts[1].strip()
    if any(c.isdigit() for c in s) or len(s) <= 8 or "." in s:
        return s, ""
    return "", s

def get_vol_sat_combined(v1, s1, v2, s2):
    v1_str = str(v1).replace(".0", "") if pd.notna(v1) else "0"
    s1_str = str(s1).strip() if pd.notna(s1) else ""
    v2_str = str(v2).replace(".0", "") if pd.notna(v2) else "0"
    s2_str = str(s2).strip() if pd.notna(s2) else ""
    
    if s2_str in ["", "-"] or v2_str == "0" or v2_str == "":
        return f"{v1_str} {s1_str}"
    return f"{v1_str} {s1_str} x {v2_str} {s2_str}"

# ==========================================
# 2. DATABASE USER & LOGIN
# ==========================================
USER_CREDENTIALS = {
    "admin": {"password": "adminfib", "role": "admin", "nama_tampil": "Fakultas Ilmu Budaya (Admin)"},
    "sasindo": {"password": "123", "role": "prodi", "nama_tampil": "Sastra Indonesia"},
    "sasing": {"password": "123", "role": "prodi", "nama_tampil": "Sastra Inggris"},
    "etno": {"password": "123", "role": "prodi", "nama_tampil": "Etnomusikologi"},
    "tari": {"password": "123", "role": "prodi", "nama_tampil": "Tari"},
    "kajian": {"password": "123", "role": "prodi", "nama_tampil": "Kajian Budaya (S2)"},
    "p2mf": {"password": "123", "role": "prodi", "nama_tampil": "Pusat Penjaminan Mutu Fakultas"}
}

if "logged_in" not in st.session_state:
    st.session_state.update({"logged_in": False, "role": None, "nama_user": None, "username": None})

if not st.session_state["logged_in"]:
    _, col_tengah, _ = st.columns([1, 2, 1])
    with col_tengah:
        st.title("🔐 Login Sistem RKA")
        st.subheader("Fakultas Ilmu Budaya - Unmul")
        with st.form("form_login"):
            u, p = st.text_input("Username"), st.text_input("Password", type="password")
            if st.form_submit_button("Masuk", type="primary"):
                if u in USER_CREDENTIALS and USER_CREDENTIALS[u]["password"] == p:
                    st.session_state.update({"logged_in": True, "role": USER_CREDENTIALS[u]["role"], "nama_user": USER_CREDENTIALS[u]["nama_tampil"], "username": u})
                    st.rerun()
                else: st.error("Username atau Password salah.")
    st.stop()

# ==========================================
# 3. SIDEBAR & MENU KHUSUS
# ==========================================
with st.sidebar:
    st.header("Sistem Perencanaan")
    st.markdown(f"👤 **{st.session_state['nama_user']}**")
    
    menu_pilihan = "1. Dashboard Kompiler Usulan"
    if st.session_state.get("role") == "admin":
        st.markdown("---")
        menu_pilihan = st.radio("📍 Navigasi Admin:", ["1. Dashboard Kompiler Usulan", "2. Pengolah Dokumen RAB"])
        st.markdown("---")
        
    if st.button("🚪 Logout", type="primary"):
        st.session_state.update({"logged_in": False, "role": None, "nama_user": None, "username": None})
        st.rerun()

# ==========================================
# FUNGSI EXPORT DATA KOMPILER (HTML)
# ==========================================
def generate_html_report(df_data, nama_prodi, hidden=False):
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        @page {{ size: A4; margin: 15mm 20mm; @bottom-right {{ content: "Halaman " counter(page) " dari " counter(pages); font-size: 9pt; color: #718096; font-family: 'Arial', sans-serif; }} }}
        *, *::before, *::after {{ box-sizing: border-box; }}
        body {{ font-family: 'Arial', sans-serif; color: #000; line-height: 1.4; margin: 0; padding: 0; font-size: 11pt; }}
        .kop-surat {{ display: flex; align-items: center; justify-content: center; border-bottom: 4px solid #000; padding-bottom: 8px; margin-bottom: 25px; position: relative; }}
        .kop-logo {{ position: absolute; left: 0; top: 5px; width: 85px; height: auto; }}
        .kop-teks {{ text-align: center; font-family: 'Times New Roman', Times, serif; flex-grow: 1; padding-left: 95px; }}
        .kop-teks h1 {{ font-size: 15pt; margin: 0; font-weight: normal; letter-spacing: 0.3px; }}
        .kop-teks h2 {{ font-size: 15pt; margin: 0; font-weight: normal; text-transform: uppercase; }}
        .kop-teks h3 {{ font-size: 17pt; margin: 0; font-weight: bold; text-transform: uppercase; }}
        .kop-teks p {{ font-size: 10.5pt; margin: 1px 0 0 0; font-family: 'Arial', sans-serif; }}
        .judul-laporan {{ text-align: center; margin-bottom: 25px; }}
        .judul-laporan h3 {{ font-size: 14pt; margin: 0 0 5px 0; text-transform: uppercase; }}
        .badge-prodi {{ display: inline-block; background-color: #f3f4f6; color: #111827; padding: 4px 12px; border-radius: 4px; font-weight: bold; border: 1px solid #d1d5db; }}
        .sub-judul-prodi {{ color: #1a365d; border-bottom: 2px solid #000; padding-bottom: 5px; margin-top: 30px; margin-bottom: 15px; font-size: 12pt; font-weight: bold; text-transform: uppercase; }}
        .block-kegiatan {{ margin-bottom: 25px; page-break-inside: avoid; }}
        .header-kegiatan {{ background-color: #f3f4f6; border: 1px solid #000; padding: 8px 12px; font-weight: bold; }}
        .catatan-review {{ font-style: italic; color: #4b5563; font-size: 10pt; padding: 4px 12px; border-left: 1px solid #000; border-right: 1px solid #000; }}
        .tabel-rincian {{ width: 100%; border-collapse: collapse; border: 1px solid #000; }}
        .tabel-rincian th {{ border: 1px solid #000; padding: 6px; font-weight: bold; text-align: center; background-color: #f9fafb; }}
        .tabel-rincian td {{ border: 1px solid #000; padding: 6px; }}
        .text-right {{ text-align: right !important; }}
        .text-center {{ text-align: center !important; }}
        .total-row td {{ font-weight: bold; background-color: #f3f4f6 !important; }}
    </style>
    </head>
    <body>
        <div class="kop-surat">
            <img src="https://lh3.googleusercontent.com/d/13kT8UkeAomtnzXVMaVRi9KWrU2IceX4r" class="kop-logo" alt="Logo Unmul">
            <div class="kop-teks">
                <h1>KEMENTERIAN PENDIDIKAN TINGGI, SAINS,</h1>
                <h1>DAN TEKNOLOGI</h1>
                <h2>UNIVERSITAS MULAWARMAN</h2>
                <h3>FAKULTAS ILMU BUDAYA</h3>
                <p>Jl. Ki Hajar Dewantara, Kampus Gunung Kelua, Samarinda 75123</p>
                <p>Telepon (0541) 7809033</p>
                <p>Laman http://fib.unmul.ac.id   Surel fib@unmul.ac.id</p>
            </div>
        </div>
        <div class="judul-laporan">
            <h3>Rekapitulasi Rincian Anggaran Tahun 2026</h3>
            <span class="badge-prodi">Program Studi: {nama_prodi}</span>
        </div>
    """
    prodi_list = sorted(df_data["Program_Studi"].unique())
    for prodi in prodi_list:
        if nama_prodi == "Seluruh Fakultas":
            html += f"<div class='sub-judul-prodi'>▶ PROGRAM STUDI: {prodi}</div>"
            
        df_p = df_data[df_data["Program_Studi"] == prodi]
        rekap_keg_prodi = df_p.groupby("Nama_Kegiatan")["Total_Usulan"].sum().reset_index()
        
        for idx, row in rekap_keg_prodi.iterrows():
            keg = row["Nama_Kegiatan"]
            tot = row["Total_Usulan"]
            df_k = df_p[df_p["Nama_Kegiatan"] == keg]
            stat = df_k["Status"].iloc[0]
            cat = df_k["Catatan_Fakultas"].iloc[0]
            teks_tot = "Rp ***" if hidden else f"Rp {format_rupiah(tot)}"
            
            html += f'<div class="block-kegiatan"><div class="header-kegiatan">{idx+1}. {keg.upper()}  |  {teks_tot}  |  [{stat}]</div>'
            if cat != "-": html += f'<div class="catatan-review"><strong>Catatan Review:</strong> {cat}</div>'
            html += '<table class="tabel-rincian"><thead><tr><th style="width: 50%; text-align: left;">Rincian Belanja</th><th style="width: 10%;">Vol</th><th style="width: 15%;">Satuan</th>'
            if not hidden: html += '<th style="width: 12%; text-align: right;">Harga Satuan</th><th style="width: 13%; text-align: right;">Total</th>'
            html += '</tr></thead><tbody>'
            
            for _, r in df_k.iterrows():
                html += f"<tr><td>{r['Rincian_Belanja']}</td><td class='text-center'>{r['Volume']}</td><td class='text-center'>{r['Satuan']}</td>"
                if not hidden: html += f"<td class='text-right'>{format_rupiah(r['Harga_Satuan'])}</td><td class='text-right'>{format_rupiah(r['Total_Usulan'])}</td>"
                html += "</tr>"
                
            if not hidden: html += f'<tr class="total-row"><td colspan="4" class="text-right">Total Usulan Anggaran Kegiatan</td><td class="text-right">{format_rupiah(tot)}</td></tr>'
            html += '</tbody></table></div>'
    html += "</body></html>"
    return html

def format_df_ke_hirarki(df_mentah, hidden=False):
    df_h = df_mentah.sort_values(by=["Program_Studi", "Nama_Kegiatan"]).copy()
    if not hidden:
        df_h["Harga_Satuan"] = df_h["Harga_Satuan"].apply(lambda x: format_rupiah(x))
        df_h["Total_Usulan"] = df_h["Total_Usulan"].apply(lambda x: format_rupiah(x))
    
    is_duplicate = df_h.duplicated(subset=["Program_Studi", "Nama_Kegiatan"])
    df_h.loc[is_duplicate, "Program_Studi"] = ""
    df_h.loc[is_duplicate, "Nama_Kegiatan"] = ""
    df_h.loc[is_duplicate, "Status"] = ""
    df_h.loc[is_duplicate, "Catatan_Fakultas"] = ""
    
    df_h = df_h[["Program_Studi", "Nama_Kegiatan", "Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan", "Total_Usulan", "Status", "Catatan_Fakultas"]]
    df_h.rename(columns={"Program_Studi": "Program Studi", "Nama_Kegiatan": "Nama Kegiatan", "Rincian_Belanja": "Rincian Belanja", "Harga_Satuan": "Harga Satuan (Rp)", "Total_Usulan": "Total Rincian (Rp)", "Catatan_Fakultas": "Catatan Review"}, inplace=True)
    if hidden: df_h.drop(columns=["Harga Satuan (Rp)", "Total Rincian (Rp)"], inplace=True)
    return df_h

def generate_excel(df_to_save, nama_sheet):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_to_save.to_excel(writer, index=False, sheet_name=nama_sheet)
    return output.getvalue()


# ==========================================
# 4. TAMPILAN PRODI
# ==========================================
if st.session_state["role"] == "prodi":
    st.title(f"📍 Panel Usulan: {st.session_state['nama_user']}")
    tab_dash, tab_baru, tab_riwayat = st.tabs(["📊 Dashboard Prodi", "📤 Buat Usulan Baru", "📜 Monitoring & Revisi"])

    with tab_dash:
        my_data = df_usulan[df_usulan["Program_Studi"] == st.session_state["nama_user"]]
        col1, col2, col3 = st.columns(3)
        total_pengajuan = my_data["Total_Usulan"].sum()
        col1.metric("Total Anggaran Diajukan", f"Rp {total_pengajuan:,.0f}".replace(',', '.'))
        col2.metric("Jumlah Kegiatan", my_data["Nama_Kegiatan"].nunique())
        col3.metric("Usulan Disetujui", len(my_data[my_data["Status"] == "Disetujui"]["Nama_Kegiatan"].unique()))
        
        st.markdown("---")
        st.markdown("### 🤖 Insight & Langkah Selanjutnya")
        insights = []
        keg_tanpa_tor = my_data[my_data["File_TOR"] == "-"]["Nama_Kegiatan"].unique()
        if len(keg_tanpa_tor) > 0: insights.append(f"❌ **Lengkapi Dokumen:** Ada {len(keg_tanpa_tor)} kegiatan belum memiliki TOR.")
        keg_rev = my_data[my_data["Status"] == "Perlu Revisi"]["Nama_Kegiatan"].unique()
        if len(keg_rev) > 0: insights.append(f"⚠️ **Tindak Lanjut Revisi:** Ada {len(keg_rev)} kegiatan perlu perbaikan.")
        
        if not insights: st.success("✅ Seluruh usulan Anda sudah lengkap dan sedang dalam proses review.")
        else:
            for item in insights: st.write(item)

        st.markdown("---")
        st.markdown("### 📋 Daftar Kegiatan yang Sudah Diinput")
        if not my_data.empty:
            rekap_keg_prodi = my_data.groupby("Nama_Kegiatan")["Total_Usulan"].sum().reset_index()
            for k in rekap_keg_prodi["Nama_Kegiatan"]:
                df_k = my_data[my_data["Nama_Kegiatan"] == k].copy()
                total_keg_val = df_k["Total_Usulan"].sum()
                status_keg = df_k["Status"].iloc[0]
                catatan_keg = df_k["Catatan_Fakultas"].iloc[0]
                
                with st.expander(f"📌 {k.upper()} | Total: Rp {total_keg_val:,.0f} | Status: {status_keg}".replace(',', '.')):
                    if catatan_keg != "-": st.warning(f"**Catatan Fakultas:** {catatan_keg}")
                    if status_keg in ["Menunggu Review", "Perlu Revisi"]:
                        col_edit1, col_edit2 = st.columns([2, 1])
                        new_nama_keg = col_edit1.text_input("Nama Kegiatan:", value=k, key=f"edit_nama_{k}")
                        prio_lama = df_k["Prioritas"].iloc[0] if "Prioritas" in df_k.columns else "Sedang"
                        if prio_lama not in ["Tinggi", "Sedang", "Rendah"]: prio_lama = "Sedang"
                        new_prio = col_edit2.selectbox("Prioritas:", ["Tinggi", "Sedang", "Rendah"], index=["Tinggi", "Sedang", "Rendah"].index(prio_lama), key=f"edit_prio_{k}")
                        
                        df_editable = df_k[["Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan"]].reset_index(drop=True)
                        df_editable.insert(0, "Hapus", False)
                        edited_keg_rows = st.data_editor(
                            df_editable, num_rows="dynamic", use_container_width=True, hide_index=True, key=f"prod_dash_ed_{k}",
                            column_config={
                                "Hapus": st.column_config.CheckboxColumn("Hapus?", default=False),
                                "Rincian_Belanja": st.column_config.TextColumn("Rincian Belanja", required=True),
                                "Volume": st.column_config.NumberColumn("Volume", min_value=0, required=True),
                                "Satuan": st.column_config.SelectboxColumn("Satuan", options=["Unit", "Orang", "Hari", "Bulan", "Tahun", "Jam", "Paket", "Stel", "Kegiatan"], required=True),
                                "Harga_Satuan": st.column_config.NumberColumn("Harga Satuan (Rp)", min_value=0, required=True)
                            }
                        )
                        c_btn1, c_btn2 = st.columns([1, 4])
                        with c_btn1:
                            if st.button("💾 Simpan Perubahan", key=f"save_prod_dash_{k}"):
                                edited_keg_rows["Volume"] = pd.to_numeric(edited_keg_rows["Volume"]).fillna(0)
                                edited_keg_rows["Harga_Satuan"] = pd.to_numeric(edited_keg_rows["Harga_Satuan"]).fillna(0)
                                edited_keg_rows["Hapus"] = edited_keg_rows["Hapus"].fillna(False)
                                edited_keg_rows["Rincian_Belanja"] = edited_keg_rows["Rincian_Belanja"].fillna("")
                                
                                valid_edited = edited_keg_rows[(edited_keg_rows["Hapus"] == False) & (edited_keg_rows["Rincian_Belanja"].str.strip() != "")]
                                df_usulan = df_usulan[~((df_usulan["Program_Studi"] == st.session_state["nama_user"]) & (df_usulan["Nama_Kegiatan"] == k))]
                                
                                if not valid_edited.empty:
                                    tgl_up = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
                                    updated_entries = pd.DataFrame([{
                                        "Tanggal_Input": tgl_up, "Program_Studi": st.session_state["nama_user"], "Nama_Kegiatan": new_nama_keg.strip(),
                                        "Rincian_Belanja": r["Rincian Belanja"], "Volume": r["Volume"], "Satuan": r["Satuan"],
                                        "Harga_Satuan": r["Harga_Satuan"], "Total_Usulan": r["Volume"] * r["Harga_Satuan"],
                                        "Prioritas": new_prio, "Status": "Menunggu Review", "Catatan_Fakultas": "-",
                                        "File_TOR": df_k["File_TOR"].iloc[0] if "File_TOR" in df_k.columns else "-"
                                    } for _, r in valid_edited.iterrows()])
                                    df_usulan = pd.concat([df_usulan, updated_entries], ignore_index=True)
                                save_data(df_usulan); st.success("Tersimpan!"); st.rerun()
                        with c_btn2:
                            if st.button("🚨 Hapus Kegiatan", key=f"del_prod_dash_{k}", type="secondary"):
                                df_usulan = df_usulan[~((df_usulan["Program_Studi"] == st.session_state["nama_user"]) & (df_usulan["Nama_Kegiatan"] == k))]
                                save_data(df_usulan); st.success("Dihapus!"); st.rerun()
                    else:
                        st.info(f"🔒 Data tidak dapat diedit karena telah menerima keputusan Fakultas (**{status_keg}**).")
                        st.dataframe(df_k[["Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan", "Total_Usulan"]].style.format({"Harga_Satuan": format_rupiah, "Total_Usulan": format_rupiah}), hide_index=True, use_container_width=True)
        else:
            st.info("Belum ada kegiatan yang diusulkan. Silakan buat usulan baru di tab 'Buat Usulan Baru'.")

    with tab_baru:
        with st.form("form_input", clear_on_submit=True):
            nama_keg = st.text_input("Nama Kegiatan Utama")
            file_tor = st.file_uploader("Upload TOR (PDF, Maks 5MB)", type=["pdf"])
            template = pd.DataFrame([{"Rincian Belanja": "", "Volume": 0, "Satuan": "Orang", "Harga Satuan": 0}])
            edited = st.data_editor(template, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Satuan": st.column_config.SelectboxColumn("Satuan", options=["Unit", "Orang", "Hari", "Bulan", "Tahun", "Jam", "Paket", "Stel", "Kegiatan"])})
            if st.form_submit_button("Kirim Usulan"):
                valid = edited[edited["Rincian Belanja"].str.strip() != ""]
                if nama_keg and not valid.empty:
                    path_tor = "-"
                    if file_tor:
                        path_tor = os.path.join(UPLOAD_DIR, f"TOR_{st.session_state['username']}_{nama_keg[:10]}.pdf")
                        with open(path_tor, "wb") as f: f.write(file_tor.getbuffer())
                    
                    new_data = pd.DataFrame([{
                        "Tanggal_Input": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"), "Program_Studi": st.session_state["nama_user"], "Nama_Kegiatan": nama_keg,
                        "Rincian_Belanja": r["Rincian Belanja"], "Volume": r["Volume"], "Satuan": r["Satuan"],
                        "Harga_Satuan": r["Harga Satuan"], "Total_Usulan": r["Volume"] * r["Harga Satuan"],
                        "Prioritas": "Sedang", "Status": "Menunggu Review", "Catatan_Fakultas": "-", "File_TOR": path_tor
                    } for _, r in valid.iterrows()])
                    df_usulan = pd.concat([df_usulan, new_data], ignore_index=True)
                    save_data(df_usulan); st.success("Terkirim!"); st.rerun()

    with tab_riwayat:
        my_data = df_usulan[df_usulan["Program_Studi"] == st.session_state["nama_user"]]
        if not my_data.empty:
            sel_keg = st.selectbox("Pilih Kegiatan untuk Direvisi/Update TOR:", my_data["Nama_Kegiatan"].unique())
            df_curr = my_data[my_data["Nama_Kegiatan"] == sel_keg]
            status_saat_ini = df_curr['Status'].iloc[0]
            st.info(f"Status: {status_saat_ini} | Catatan: {df_curr['Catatan_Fakultas'].iloc[0]}")
            
            if status_saat_ini in ["Menunggu Review", "Perlu Revisi"]:
                df_to_edit = df_curr[["Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan"]].reset_index(drop=True)
                df_to_edit.insert(0, "Hapus", False)
                rev_ed = st.data_editor(df_to_edit, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Hapus": st.column_config.CheckboxColumn("Hapus?", default=False), "Satuan": st.column_config.SelectboxColumn("Satuan", options=["Unit", "Orang", "Hari", "Bulan", "Tahun", "Jam", "Paket", "Stel", "Kegiatan"])})
                if st.button("Kirim Ulang Revisi"):
                    rev_ed["Volume"] = pd.to_numeric(rev_ed["Volume"]).fillna(0)
                    rev_ed["Harga_Satuan"] = pd.to_numeric(rev_ed["Harga_Satuan"]).fillna(0)
                    rev_ed["Hapus"] = rev_ed["Hapus"].fillna(False)
                    rev_ed["Rincian_Belanja"] = rev_ed["Rincian_Belanja"].fillna("")
                    valid_rev = rev_ed[(rev_ed["Hapus"] == False) & (rev_ed["Rincian_Belanja"].str.strip() != "")]
                    df_usulan = df_usulan[~((df_usulan["Program_Studi"]==st.session_state["nama_user"]) & (df_usulan["Nama_Kegiatan"]==sel_keg))]
                    if not valid_rev.empty:
                        rev_entries = pd.DataFrame([{
                            "Tanggal_Input": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"), "Program_Studi": st.session_state["nama_user"], "Nama_Kegiatan": sel_keg,
                            "Rincian_Belanja": r["Rincian_Belanja"], "Volume": r["Volume"], "Satuan": r["Satuan"],
                            "Harga_Satuan": r["Harga_Satuan"], "Total_Usulan": r["Volume"] * r["Harga_Satuan"],
                            "Prioritas": df_curr["Prioritas"].iloc[0], "Status": "Menunggu Review", "Catatan_Fakultas": f"Revisi: {df_curr['Catatan_Fakultas'].iloc[0]}", "File_TOR": df_curr["File_TOR"].iloc[0]
                        } for _, r in valid_rev.iterrows()])
                        df_usulan = pd.concat([df_usulan, rev_entries], ignore_index=True)
                    save_data(df_usulan); st.success("Revisi dikirim!"); st.rerun()
            else:
                st.table(df_curr[["Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan", "Total_Usulan"]].style.format({"Harga_Satuan": format_rupiah, "Total_Usulan": format_rupiah}))
            
            with st.expander("📄 Update / Susulan Dokumen TOR"):
                new_tor = st.file_uploader("Upload PDF (Maks 5MB)", type=["pdf"], key=f"up_{sel_keg}")
                if st.button("Simpan Dokumen", key=f"btn_{sel_keg}"):
                    if new_tor:
                        safe_name = "".join([c for c in sel_keg if c.isalnum()]).rstrip()
                        path = os.path.join(UPLOAD_DIR, f"TOR_UPD_{st.session_state['username']}_{safe_name}.pdf")
                        with open(path, "wb") as f: f.write(new_tor.getbuffer())
                        df_usulan.loc[(df_usulan["Program_Studi"]==st.session_state["nama_user"]) & (df_usulan["Nama_Kegiatan"]==sel_keg), "File_TOR"] = path
                        save_data(df_usulan); st.success("TOR Terupdate!"); st.rerun()

# ==========================================
# 5. TAMPILAN ADMIN
# ==========================================
elif st.session_state["role"] == "admin":
    
    if menu_pilihan == "1. Dashboard Kompiler Usulan":
        st.title("📊 Dashboard Monitoring & Review")
        
        col_info, col_toggle = st.columns([3, 1])
        with col_info: st.info("Gunakan tab di bawah untuk meninjau usulan dari setiap Program Studi.")
        with col_toggle: sembunyikan_nilai = st.toggle("🙈 Sembunyikan Nominal Anggaran", value=False)
        
        tab_rev, tab_hapus, tab_ins, tab_restore = st.tabs(["📋 Review & Analisis", "🗑️ Manajemen Data", "🤖 Insight", "♻️ Restore Data"])

        with tab_restore:
            st.subheader("♻️ Restore Database dari File Cadangan")
            st.warning("Upload file **Rekap_FIB.csv** (atau file Excel Backup Anda) di bawah ini untuk mengembalikan seluruh data prodi yang hilang.")
            
            file_cadangan = st.file_uploader("Upload File Backup CSV/Excel Anda di sini", type=["csv", "xlsx"])
            if st.button("🚀 Jalankan Restore Data", type="primary"):
                if file_cadangan is not None:
                    try:
                        if file_cadangan.name.endswith('.csv'): df_pulih = pd.read_csv(file_cadangan)
                        else: df_pulih = pd.read_excel(file_cadangan)
                        
                        kolom_wajib = ["Program_Studi", "Nama_Kegiatan", "Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan", "Total_Usulan"]
                        if all(k in df_pulih.columns for k in kolom_wajib):
                            if "Tanggal_Input" not in df_pulih.columns: df_pulih["Tanggal_Input"] = "-"
                            if "Prioritas" not in df_pulih.columns: df_pulih["Prioritas"] = "Sedang"
                            if "Status" not in df_pulih.columns: df_pulih["Status"] = "Menunggu Review"
                            if "Catatan_Fakultas" not in df_pulih.columns: df_pulih["Catatan_Fakultas"] = "-"
                            if "File_TOR" not in df_pulih.columns: df_pulih["File_TOR"] = "-"
                            save_data(df_pulih); st.success("🎉 Seluruh data telah berhasil dipulihkan!"); st.rerun()
                        else: st.error("Gagal! Format kolom pada file yang diupload tidak cocok dengan database aplikasi.")
                    except Exception as e: st.error(f"Kesalahan saat membaca file: {e}")
                else: st.error("Pilih file CSV atau Excel terlebih dahulu.")

        with tab_rev:
            if df_usulan.empty: st.warning("Data kosong. Silakan gunakan tab 'Restore Data' jika Anda memiliki backup.")
            else:
                st.subheader("🏙️ Rekapitulasi Anggaran Per Prodi")
                rekap_semua = df_usulan.groupby("Program_Studi")["Total_Usulan"].sum().reset_index()
                rekap_semua.columns = ["Program Studi", "Total Usulan (Rp)"]
                if sembunyikan_nilai:
                    rekap_semua["Total Usulan (Rp)"] = "Rp ***"
                    st.table(rekap_semua)
                else: st.table(rekap_semua.style.format({"Total Usulan (Rp)": format_rupiah}))

                st.markdown("---")
                prodi_sel = st.selectbox("Pilih Prodi untuk melihat daftar kegiatan:", sorted(df_usulan["Program_Studi"].unique()))
                df_p = df_usulan[df_usulan["Program_Studi"] == prodi_sel]
                rekap_keg_prodi = df_p.groupby("Nama_Kegiatan")["Total_Usulan"].sum().reset_index()
                
                for k in rekap_keg_prodi["Nama_Kegiatan"]:
                    df_k = df_p[df_p["Nama_Kegiatan"] == k].copy()
                    total_keg_val = df_k["Total_Usulan"].sum()
                    teks_nominal_expander = "Rp ***" if sembunyikan_nilai else f"Rp {format_rupiah(total_keg_val)}"
                    
                    with st.expander(f"📌 {k.upper()} | Total Usulan: {teks_nominal_expander} | Status: {df_k['Status'].iloc[0]}"):
                        path = df_k["File_TOR"].iloc[0]
                        if path != "-" and os.path.exists(path):
                            with open(path, "rb") as f: st.download_button("📥 Download TOR", f, file_name=os.path.basename(path), key=f"dl_{prodi_sel}_{k}")
                        
                        c1, c2 = st.columns([1, 2])
                        n_s = c1.selectbox("Update Status:", ["Menunggu Review", "Disetujui", "Perlu Revisi", "Ditolak"], index=["Menunggu Review", "Disetujui", "Perlu Revisi", "Ditolak"].index(df_k["Status"].iloc[0]), key=f"s_{prodi_sel}_{k}")
                        n_n = c2.text_area("Catatan Fakultas:", value=df_k["Catatan_Fakultas"].iloc[0], key=f"n_{prodi_sel}_{k}")
                        
                        if st.button("Simpan Hasil Review", key=f"b_{prodi_sel}_{k}"):
                            df_usulan.loc[(df_usulan["Program_Studi"]==prodi_sel) & (df_usulan["Nama_Kegiatan"]==k), ["Status", "Catatan_Fakultas"]] = [n_s, n_n]
                            save_data(df_usulan); st.success(f"Status diperbarui!"); st.rerun()
                        
                        if sembunyikan_nilai: st.dataframe(df_k[["Rincian_Belanja", "Volume", "Satuan"]].copy(), hide_index=True, use_container_width=True)
                        else: st.dataframe(df_k[["Rincian_Belanja", "Volume", "Satuan", "Harga_Satuan", "Total_Usulan"]].style.format({"Harga_Satuan": format_rupiah, "Total_Usulan": format_rupiah}), hide_index=True, use_container_width=True)

        with tab_hapus:
            st.subheader("🗑️ Hapus Data Rincian")
            if not df_usulan.empty:
                opsi_hapus = {idx: f"[{row['Program_Studi']}] {row['Nama_Kegiatan']} - {row['Rincian_Belanja']}" for idx, row in df_usulan.iterrows()}
                sel_h = st.selectbox("Pilih data rincian belanja:", options=list(opsi_hapus.keys()), format_func=lambda x: opsi_hapus[x])
                if st.button("🚨 Hapus Permanen", type="primary"):
                    df_usulan = df_usulan.drop(index=sel_h).reset_index(drop=True)
                    save_data(df_usulan); st.success("Dihapus!"); st.rerun()

        with tab_ins:
            st.subheader("🤖 Analisis & Insight Pintar")
            if not df_usulan.empty:
                if sembunyikan_nilai: st.warning("⚠️ Insight dinonaktifkan di mode sembunyi.")
                else:
                    tot = df_usulan['Total_Usulan'].sum()
                    if tot > 0:
                        st.info(f"💡 **Total anggaran diajukan: Rp {format_rupiah(tot)}**")
                        st.bar_chart(df_usulan.groupby("Program_Studi")["Total_Usulan"].sum().reset_index().set_index("Program_Studi")["Total_Usulan"])
                        status_pivot = pd.crosstab(df_usulan["Program_Studi"], df_usulan["Status"]).reset_index()
                        rekap_detail = df_usulan.groupby("Program_Studi").agg(Total_Anggaran=("Total_Usulan", "sum"), Jumlah_Kegiatan=("Nama_Kegiatan", "nunique")).reset_index()
                        rekap_final = pd.merge(rekap_detail, status_pivot, on="Program_Studi", how="left").fillna(0)
                        for stat in ["Menunggu Review", "Disetujui", "Perlu Revisi", "Ditolak"]:
                            if stat not in rekap_final.columns: rekap_final[stat] = 0
                        rekap_final.rename(columns={"Program_Studi": "Program Studi", "Jumlah_Kegiatan": "Jml Kegiatan", "Total_Anggaran": "Total Anggaran (Rp)"}, inplace=True)
                        st.dataframe(rekap_final[["Program Studi", "Jml Kegiatan", "Menunggu Review", "Disetujui", "Perlu Revisi", "Ditolak", "Total Anggaran (Rp)"]].style.format({"Total Anggaran (Rp)": format_rupiah}), hide_index=True, use_container_width=True)

                st.markdown("---")
                st.markdown("### 📄 Laporan Cetak Hirarki & PDF Print-Ready")
                prodi_ins_sel = st.selectbox("Pilih Program Studi untuk dilihat:", sorted(df_usulan["Program_Studi"].unique()), key="ins_prodi_sel")
                df_ins_p = df_usulan[df_usulan["Program_Studi"] == prodi_ins_sel]
                
                if not df_ins_p.empty:
                    st.dataframe(format_df_ke_hirarki(df_ins_p, hidden=sembunyikan_nilai), hide_index=True, use_container_width=True)
                    
                    col_ex1, col_ex2 = st.columns(2)
                    with col_ex1: st.download_button("📊 Excel: Laporan Prodi", data=generate_excel(format_df_ke_hirarki(df_ins_p, hidden=sembunyikan_nilai), prodi_ins_sel[:30]), file_name=f"Laporan_{prodi_ins_sel}_2026.xlsx", use_container_width=True)
                    with col_ex2: st.download_button("📊 Excel: Laporan Fakultas", data=generate_excel(format_df_ke_hirarki(df_usulan, hidden=sembunyikan_nilai), "Seluruh_Fakultas"), file_name="Laporan_FIB_Semua_2026.xlsx", use_container_width=True)
                    
                    col_pdf1, col_pdf2 = st.columns(2)
                    with col_pdf1: st.download_button("📑 PDF: Laporan Prodi (Web)", data=generate_html_report(df_ins_p, prodi_ins_sel, hidden=sembunyikan_nilai).encode('utf-8'), file_name=f"Cetak_{prodi_ins_sel}.html", mime="text/html", help="Tekan Ctrl+P di browser.", use_container_width=True)
                    with col_pdf2: st.download_button("📑 PDF: Laporan Fakultas (Web)", data=generate_html_report(df_usulan, "Seluruh Fakultas", hidden=sembunyikan_nilai).encode('utf-8'), file_name="Cetak_FIB_Semua.html", mime="text/html", help="Tekan Ctrl+P di browser.", use_container_width=True)

    # ----------------------------------------------------
    # MENU 2: PENGOLAH RAB (FINAL COMPACT VIEW + AUTO RESTORE)
    # ----------------------------------------------------
    elif menu_pilihan == "2. Pengolah Dokumen RAB":
        st.title("📄 Pengolah Dokumen RAB Universitas")
        st.caption("Sistem Manajemen & Generator RAB Berjenjang dengan Pemisahan Kode Otomatis.")

        tab_master, tab_buat, tab_daftar = st.tabs(["🗂️ Master Database", "📝 Buat RAB Baru", "📂 Daftar RAB Tersimpan"])

        with tab_master:
            st.info("💡 Input Master Data. Format bebas, mesin otomatis memisahkan teks sebelum tanda strip '-' ke kolom Kode Excel.")
            
            with st.expander("⚡ Restore Database Master FIB (Otomatis)", expanded=True):
                st.warning("Klik tombol di bawah ini untuk memulihkan seluruh data standar KRO, RO, Komponen, dan 50+ Akun Belanja FIB jika data Anda hilang akibat server restart.")
                if st.button("🚀 Restore Data Standar FIB", type="primary"):
                    df_kro_baru = pd.DataFrame({"KRO": ["7729.BEI - Bantuan Lembaga", "7730.CAA - Sarana Bidang Pendidikan", "7730.CBJ - Prasarana Bidang Pendidikan Tinggi", "7730.DBA - Pendidikan Tinggi"]})
                    save_table(df_kro_baru, "rab_m_kro")
                    
                    df_ro_baru = pd.DataFrame([
                        {"KRO": "7729.BEI - Bantuan Lembaga", "RO": "7729.BEI.001 - PT Penerima Bantuan Dukungan Operasional"},
                        {"KRO": "7729.BEI - Bantuan Lembaga", "RO": "7729.BEI.002 - PT Penerima Bantuan Pembelajaran"},
                        {"KRO": "7729.BEI - Bantuan Lembaga", "RO": "7729.BEI.004 - PT Penerima Bantuan Sarana dan Prasarana Pembelajaran"},
                        {"KRO": "7730.CAA - Sarana Bidang Pendidikan", "RO": "7730.CAA.001 - Sarana Pendukung Pembelajaran"},
                        {"KRO": "7730.CAA - Sarana Bidang Pendidikan", "RO": "7730.CAA.002 - Sarana Pendukung Perkantoran"},
                        {"KRO": "7730.CBJ - Prasarana Bidang Pendidikan Tinggi", "RO": "7730.CBJ.001 - Prasarana Pendukung Pembelajaran"},
                        {"KRO": "7730.CBJ - Prasarana Bidang Pendidikan Tinggi", "RO": "7730.CBJ.002 - Prasarana Pendukung Perkantoran"},
                        {"KRO": "7730.DBA - Pendidikan Tinggi", "RO": "7730.DBA.001 - Layanan Pendidikan"},
                        {"KRO": "7730.DBA - Pendidikan Tinggi", "RO": "7730.DBA.002 - Dukungan Operasional Pembelajaran"},
                        {"KRO": "7730.DBA - Pendidikan Tinggi", "RO": "7730.DBA.003 - Penelitian dan Pengabdian Masyarakat"},
                        {"KRO": "7730.DBA - Pendidikan Tinggi", "RO": "7730.DBA.004 - Pengabdian Kepada Masyarakat"}
                    ])
                    save_table(df_ro_baru, "rab_m_ro")
                    
                    df_komp_baru = pd.DataFrame([
                        {"RO": "7729.BEI.001 - PT Penerima Bantuan Dukungan Operasional", "Komponen": "004 - Dukungan Operasional Penyelenggaraan Pendidikan"},
                        {"RO": "7729.BEI.002 - PT Penerima Bantuan Pembelajaran", "Komponen": "004 - Dukungan Operasional Penyelenggaraan Pendidikan"},
                        {"RO": "7729.BEI.004 - PT Penerima Bantuan Sarana dan Prasarana Pembelajaran", "Komponen": "004 - Dukungan Operasional Penyelenggaraan Pendidikan"},
                        {"RO": "7730.CAA.001 - Sarana Pendukung Pembelajaran", "Komponen": "051 - Pengadaan Sarana Pendukung Pembelajaran"},
                        {"RO": "7730.CAA.002 - Sarana Pendukung Perkantoran", "Komponen": "051 - Sarana Pendukung Perkantoran"},
                        {"RO": "7730.CBJ.001 - Prasarana Pendukung Pembelajaran", "Komponen": "051 - Pengadaan Prasarana Pendukung Pembelajaran"},
                        {"RO": "7730.CBJ.002 - Prasarana Pendukung Perkantoran", "Komponen": "051 - Pengadaan Prasarana Pendukung Perkantoran"},
                        {"RO": "7730.DBA.001 - Layanan Pendidikan", "Komponen": "051 - Pemeliharaan Sarana dan Prasarana Pembelajaran"},
                        {"RO": "7730.DBA.001 - Layanan Pendidikan", "Komponen": "052 - Pemeliharaan Sarana dan Prasarana Perkantoran"},
                        {"RO": "7730.DBA.001 - Layanan Pendidikan", "Komponen": "053 - Penyelenggaraan Layanan Pendidikan Perguruan Tinggi"},
                        {"RO": "7730.DBA.002 - Dukungan Operasional Pembelajaran", "Komponen": "051 - Penyelenggaraan Dukungan Operasional Pembelajaran"},
                        {"RO": "7730.DBA.002 - Dukungan Operasional Pembelajaran", "Komponen": "053 - Pelaksanaan Layanan Pengembangan Sistem Tata Kelola"},
                        {"RO": "7730.DBA.003 - Penelitian dan Pengabdian Masyarakat", "Komponen": "051 - Penelitian"},
                        {"RO": "7730.DBA.003 - Penelitian dan Pengabdian Masyarakat", "Komponen": "052 - Pengabdian Kepada Masyarakat"}
                    ])
                    save_table(df_komp_baru, "rab_m_komp")
                    
                    df_akun_baru = pd.DataFrame([
                        {"Account_Code": "521111", "Account_Name": "Belanja Keperluan Perkantoran"},
                        {"Account_Code": "521115", "Account_Name": "Belanja Honor Operasional Satuan Kerja"},
                        {"Account_Code": "521119", "Account_Name": "Belanja Barang Operasional Lainnya"},
                        {"Account_Code": "521211", "Account_Name": "Belanja Bahan"},
                        {"Account_Code": "521219", "Account_Name": "Belanja Barang Non Operasional Lainnya"},
                        {"Account_Code": "521253", "Account_Name": "Belanja Gedung dan Bangunan Ekstrakomptabel"},
                        {"Account_Code": "522111", "Account_Name": "Belanja Langganan Listrik"},
                        {"Account_Code": "522112", "Account_Name": "Belanja Langganan Telepon"},
                        {"Account_Code": "522113", "Account_Name": "Belanja Langganan Air"},
                        {"Account_Code": "522119", "Account_Name": "Belanja Langganan Daya dan Jasa Lainnya"},
                        {"Account_Code": "522121", "Account_Name": "Belanja Jasa Pos dan Giro"},
                        {"Account_Code": "522131", "Account_Name": "Belanja Jasa Konsultan"},
                        {"Account_Code": "522141", "Account_Name": "Belanja Sewa"},
                        {"Account_Code": "522151", "Account_Name": "Belanja Jasa Profesi"},
                        {"Account_Code": "523119", "Account_Name": "Belanja Biaya Pemeliharaan Gedung Lainnya"},
                        {"Account_Code": "523121", "Account_Name": "Belanja Biaya Pemeliharaan Peralatan dan Mesin"},
                        {"Account_Code": "523129", "Account_Name": "Belanja Biaya Pemeliharaan Peralatan Lainnya"},
                        {"Account_Code": "523131", "Account_Name": "Belanja Biaya Pemeliharaan Gedung dan Bangunan"},
                        {"Account_Code": "523132", "Account_Name": "Belanja Biaya Pemeliharaan Irigasi"},
                        {"Account_Code": "523133", "Account_Name": "Belanja Biaya Pemeliharaan Jaringan"},
                        {"Account_Code": "523199", "Account_Name": "Belanja Biaya Pemeliharaan Lainnya"},
                        {"Account_Code": "524111", "Account_Name": "Belanja Perjalanan Dinas Biasa"},
                        {"Account_Code": "524114", "Account_Name": "Belanja Perjalanan Dinas Paket Meeting Dalam Kota"},
                        {"Account_Code": "524119", "Account_Name": "Belanja Perjalanan Dinas Paket Luar Kota"},
                        {"Account_Code": "524211", "Account_Name": "Belanja Perjalanan Biasa - Luar Negeri"},
                        {"Account_Code": "525111", "Account_Name": "Belanja Gaji dan Tunjangan"},
                        {"Account_Code": "525112", "Account_Name": "Belanja Barang"},
                        {"Account_Code": "525113", "Account_Name": "Belanja Jasa"},
                        {"Account_Code": "525114", "Account_Name": "Belanja Pemeliharaan"},
                        {"Account_Code": "525115", "Account_Name": "Belanja Perjalanan Dinas"},
                        {"Account_Code": "525119", "Account_Name": "Belanja Penyediaan Barang dan Jasa Lainnya"},
                        {"Account_Code": "525162", "Account_Name": "Belanja Peralatan dan Mesin Ekstrakomptabel BLU"},
                        {"Account_Code": "525163", "Account_Name": "Belanja Gedung dan Bangunan - Ekstrakomptabel BLU"},
                        {"Account_Code": "532111", "Account_Name": "Belanja Modal Peralatan dan Mesin"},
                        {"Account_Code": "532114", "Account_Name": "Belanja Modal Sewa Peralatan dan Mesin"},
                        {"Account_Code": "532121", "Account_Name": "Belanja Penambahan Nilai Peralatan dan Mesin"},
                        {"Account_Code": "533114", "Account_Name": "Belanja Modal Sewa Peralatan Gedung"},
                        {"Account_Code": "533115", "Account_Name": "Belanja Modal Perencanaan Gedung"},
                        {"Account_Code": "533116", "Account_Name": "Belanja Modal Perizinan Gedung"},
                        {"Account_Code": "533117", "Account_Name": "Belanja Modal Pengosongan Bangunan"},
                        {"Account_Code": "533121", "Account_Name": "Belanja Penambahan Nilai Gedung"},
                        {"Account_Code": "534111", "Account_Name": "Belanja Modal Jalan dan Jembatan"},
                        {"Account_Code": "534112", "Account_Name": "Belanja Modal Bahan Baku Jalan"},
                        {"Account_Code": "534115", "Account_Name": "Belanja Modal Perencanaan Jalan"},
                        {"Account_Code": "534121", "Account_Name": "Belanja Modal Irigasi"},
                        {"Account_Code": "534125", "Account_Name": "Belanja Modal Perencanaan Irigasi"},
                        {"Account_Code": "534131", "Account_Name": "Belanja Modal Jaringan"},
                        {"Account_Code": "537112", "Account_Name": "Belanja Modal Peralatan dan Mesin - BLU"},
                        {"Account_Code": "537113", "Account_Name": "Belanja Modal Gedung dan Bangunan - BLU"},
                        {"Account_Code": "537114", "Account_Name": "Belanja Modal Jalan, Irigasi dan Jaringan - BLU"},
                        {"Account_Code": "537115", "Account_Name": "Belanja Modal Lainnya - BLU"},
                        {"Account_Code": "543122", "Account_Name": "Belanja Modal Bahan Baku Irigasi"}
                    ])
                    save_table(df_akun_baru, "rab_m_akun")
                    st.success("🎉 BOOM! Seluruh Data Master FIB berhasil dipulihkan secara otomatis!"); st.rerun()

            col_m1, col_m2 = st.columns(2)
            with col_m1:
                st.markdown("**1. Master KRO**")
                edit_kro = st.data_editor(df_m_kro, num_rows="dynamic", use_container_width=True, hide_index=True, key="me_kro")
                if st.button("💾 Simpan KRO"): save_table(edit_kro.dropna(how='all'), "rab_m_kro"); st.rerun()
                    
                st.markdown("**3. Master Komponen**")
                list_ro = df_m_ro["RO"].tolist() if not df_m_ro.empty else ["Isi Master RO Dulu"]
                edit_komp = st.data_editor(df_m_komp, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"RO": st.column_config.SelectboxColumn("Induk RO", options=list_ro, required=True)}, key="me_komp")
                if st.button("💾 Simpan Komponen"): save_table(edit_komp.dropna(how='all'), "rab_m_komp"); st.rerun()

                st.markdown("**5. Master Akun Belanja**")
                edit_akun = st.data_editor(df_m_akun, num_rows="dynamic", use_container_width=True, hide_index=True, key="me_akun")
                if st.button("💾 Simpan Akun Belanja"): save_table(edit_akun.dropna(how='all'), "rab_m_akun"); st.rerun()

            with col_m2:
                st.markdown("**2. Master RO (Rincian Output)**")
                list_kro = df_m_kro["KRO"].tolist() if not df_m_kro.empty else ["Isi Master KRO Dulu"]
                edit_ro = st.data_editor(df_m_ro, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"KRO": st.column_config.SelectboxColumn("Induk KRO", options=list_kro, required=True)}, key="me_ro")
                if st.button("💾 Simpan RO"): save_table(edit_ro.dropna(how='all'), "rab_m_ro"); st.rerun()
                
                st.markdown("**4. Master Sub-Komponen**")
                list_komp = df_m_komp["Komponen"].tolist() if not df_m_komp.empty else ["Isi Master Komponen Dulu"]
                edit_subkomp = st.data_editor(df_m_subkomp, num_rows="dynamic", use_container_width=True, hide_index=True, column_config={"Komponen": st.column_config.SelectboxColumn("Induk Komponen", options=list_komp, required=True)}, key="me_subkomp")
                if st.button("💾 Simpan Sub-Komponen"): save_table(edit_subkomp.dropna(how='all'), "rab_m_subkomp"); st.rerun()

                st.markdown("**6. Master Pejabat (Penandatangan)**")
                edit_pejabat = st.data_editor(df_m_pejabat, num_rows="dynamic", use_container_width=True, hide_index=True, key="me_pej")
                if st.button("💾 Simpan Data Pejabat"): save_table(edit_pejabat.dropna(how='all'), "rab_m_pejabat"); st.rerun()

        with tab_buat:
            if df_m_kro.empty or df_m_ro.empty or df_m_komp.empty or df_m_akun.empty:
                st.warning("⚠️ Master Database masih kosong! Isi data di tab Master terlebih dahulu.")
            else:
                st.subheader("1. Klasifikasi Output RAB")
                col_c1, col_c2 = st.columns(2)
                pilih_kro = col_c1.selectbox("Pilih KRO", df_m_kro["KRO"].tolist())
                opsi_ro = df_m_ro[df_m_ro["KRO"] == pilih_kro]["RO"].tolist()
                pilih_ro = col_c2.selectbox("Pilih RO", opsi_ro if opsi_ro else ["Tidak ada RO"])
                
                col_c3, col_c4 = st.columns(2)
                opsi_komp = df_m_komp[df_m_komp["RO"] == pilih_ro]["Komponen"].tolist()
                pilih_komp = col_c3.selectbox("Pilih Komponen", opsi_komp if opsi_komp else ["Tidak ada Komponen"])
                opsi_subkomp = df_m_subkomp[df_m_subkomp["Komponen"] == pilih_komp]["Sub_Komponen"].tolist()
                pilih_subkomp = col_c4.selectbox("Pilih Sub-Komponen", opsi_subkomp if opsi_subkomp else ["Tidak Ada Sub-Komponen"])

                st.markdown("---")
                st.subheader("2. Informasi Utama Kegiatan")
                col_u1, col_u2 = st.columns(2)
                rab_kegiatan = col_u1.text_input("Nama Kegiatan", placeholder="Contoh: Pemeliharaan Alat Operasional Pendukung TIK")
                
                _, kro_narasi = split_kode(pilih_kro) if pilih_kro else ("", "")
                kro_narasi_bersih = kro_narasi.strip("() ")
                default_sasaran = f"Peningkatan {kro_narasi_bersih}" if kro_narasi_bersih else ""
                
                rab_sasaran = col_u2.text_input("Sasaran Kegiatan", value=default_sasaran)
                rab_vol = col_u1.number_input("Volume Target", value=1, min_value=1)
                rab_satuan = col_u2.text_input("Satuan Ukur", placeholder="Contoh: Layanan / Bulan")
                
                rab_tahun = col_u1.text_input("Tahun Anggaran", value="2027")

                st.markdown("---")
                st.subheader("3. Rincian Belanja (Pengali Volume & Satuan)")
                
                opsi_akun = []
                if not df_m_akun.empty:
                    for _, row in df_m_akun.iterrows():
                        opsi_akun.append(f"{row['Account_Code']} - {row['Account_Name']}")
                
                template_detail = pd.DataFrame([{"Akun Belanja": opsi_akun[0] if opsi_akun else "", "Uraian Belanja": "", "Vol 1": 1, "Sat 1": "Unit", "Vol 2": 1, "Sat 2": "-", "Harga Satuan": 0}])
                
                df_input_detail = st.data_editor(
                    template_detail, num_rows="dynamic", use_container_width=True, hide_index=True, key="grid_buat_rab",
                    column_config={
                        "Akun Belanja": st.column_config.SelectboxColumn("Akun Belanja", options=opsi_akun, required=True),
                        "Uraian Belanja": st.column_config.TextColumn("Detail / Uraian", required=True),
                        "Vol 1": st.column_config.NumberColumn("Vol 1", min_value=1, required=True),
                        "Sat 1": st.column_config.TextColumn("Sat 1", required=True),
                        "Vol 2": st.column_config.NumberColumn("Vol 2", min_value=0),
                        "Sat 2": st.column_config.TextColumn("Sat 2 (Biarkan '-' jika tak ada)"),
                        "Harga Satuan": st.column_config.NumberColumn("Harga Satuan (Rp)", min_value=0, required=True)
                    }
                )

                df_input_detail["Vol_1_Num"] = pd.to_numeric(df_input_detail["Vol 1"]).fillna(1)
                df_input_detail["Vol_2_Num"] = pd.to_numeric(df_input_detail["Vol 2"]).fillna(1)
                df_input_detail.loc[df_input_detail["Vol_2_Num"] == 0, "Vol_2_Num"] = 1 # Hindari kali 0
                df_input_detail["Harga_Num"] = pd.to_numeric(df_input_detail["Harga Satuan"]).fillna(0)
                
                total_rab_live = (df_input_detail["Vol_1_Num"] * df_input_detail["Vol_2_Num"] * df_input_detail["Harga_Num"]).sum()
                
                st.markdown("#### 💰 Akumulasi Anggaran Alokasi Dana")
                st.metric("Total Alokasi Dana (Otomatis Mengikuti Detail)", f"Rp {format_rupiah(total_rab_live)}")
                rab_alokasi = total_rab_live

                st.markdown("---")
                st.subheader("4. Pengesahan (Penandatangan Dokumen)")
                col_p1, col_p2 = st.columns(2)
                opsi_pejabat = {idx: f"{row['Jabatan']} - {row['Nama']}" for idx, row in df_m_pejabat.iterrows()}
                pilih_pejabat = col_p1.selectbox("Pilih Pejabat Penandatangan", options=list(opsi_pejabat.keys()), format_func=lambda x: opsi_pejabat[x]) if opsi_pejabat else None
                tgl_cetak = col_p2.date_input("Tanggal Dokumen Cetak")
                
                if st.button("💾 Simpan & Terbitkan RAB", type="primary"):
                    valid_detail = df_input_detail[df_input_detail["Uraian Belanja"].str.strip() != ""].copy()
                    if not rab_kegiatan or valid_detail.empty or pilih_pejabat is None:
                        st.error("Gagal! Pastikan Nama Kegiatan, Rincian Item Belanja, dan Master Pejabat sudah lengkap.")
                    else:
                        id_rab_baru = f"RAB-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                        tgl_sekarang = datetime.now().strftime('%Y-%m-%d %H:%M')
                        dt_pjb = df_m_pejabat.loc[pilih_pejabat]
                        
                        new_utama = pd.DataFrame([{
                            "ID_RAB": id_rab_baru, "Tanggal": tgl_sekarang, "Tahun": str(rab_tahun), "Tgl_Cetak": str(tgl_cetak),
                            "KRO": pilih_kro, "RO": pilih_ro, "Komponen": pilih_komp, "Sub_Komponen": pilih_subkomp,
                            "Kegiatan": rab_kegiatan, "Sasaran": rab_sasaran, "Volume": rab_vol, "Satuan": rab_satuan, "Alokasi": rab_alokasi,
                            "Jabatan": dt_pjb['Jabatan'], "Nama_Pejabat": dt_pjb['Nama'], "NIP_Pejabat": dt_pjb['NIP']
                        }])
                        df_rab_utama = pd.concat([df_rab_utama, new_utama], ignore_index=True)
                        save_table(df_rab_utama, "rab_utama")
                        
                        valid_detail["ID_RAB"] = id_rab_baru
                        valid_detail["Total_Biaya"] = valid_detail["Vol_1_Num"] * valid_detail["Vol_2_Num"] * valid_detail["Harga_Num"]
                        valid_detail.rename(columns={"Akun Belanja": "Akun_Belanja", "Uraian Belanja": "Uraian", "Vol 1":"Vol_1", "Sat 1":"Sat_1", "Vol 2":"Vol_2", "Sat 2":"Sat_2", "Harga Satuan": "Harga_Satuan"}, inplace=True)
                        
                        df_rab_detail = pd.concat([df_rab_detail, valid_detail[["ID_RAB", "Akun_Belanja", "Uraian", "Vol_1", "Sat_1", "Vol_2", "Sat_2", "Harga_Satuan", "Total_Biaya"]]], ignore_index=True)
                        save_table(df_rab_detail, "rab_detail")
                        st.success(f"✅ RAB Resmi '{rab_kegiatan}' Berhasil Terbit!"); st.rerun()

        with tab_daftar:
            if df_rab_utama.empty: st.info("Belum ada dokumen RAB yang tersimpan.")
            else:
                st.subheader("Arsip Dokumen RAB")
                opsi_arsip = {row['ID_RAB']: f"[{row['Tanggal']}] {row['Kegiatan']}" for _, row in df_rab_utama.iterrows()}
                pilih_arsip = st.selectbox("Pilih RAB yang ingin dilihat/diunduh:", options=list(opsi_arsip.keys()), format_func=lambda x: opsi_arsip[x])
                
                head_terpilih = df_rab_utama[df_rab_utama["ID_RAB"] == pilih_arsip]
                detail_terpilih = df_rab_detail[df_rab_detail["ID_RAB"] == pilih_arsip]
                
                tahun_rab = head_terpilih.get('Tahun', pd.Series(['2027'])).iloc[0]
                if tahun_rab == "-": tahun_rab = "2027"

                df_view = detail_terpilih.copy()
                df_view['Kode Akun'] = df_view['Akun_Belanja'].apply(lambda x: split_kode(x)[0])
                df_view['Nama Akun Belanja'] = df_view['Akun_Belanja'].apply(lambda x: split_kode(x)[1])
                df_view['Volume & Satuan'] = df_view.apply(lambda r: get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2']), axis=1)
                
                st.markdown(f"**Klasifikasi Dokumen:** {head_terpilih['KRO'].iloc[0]} ➔ {head_terpilih['RO'].iloc[0]} ➔ {head_terpilih['Komponen'].iloc[0]}")
                st.markdown(f"**Total Anggaran Terakumulasi:** Rp {format_rupiah(detail_terpilih['Total_Biaya'].sum())}")
                
                st.dataframe(df_view[["Kode Akun", "Nama Akun Belanja", "Uraian", "Volume & Satuan", "Harga_Satuan", "Total_Biaya"]].style.format({"Harga_Satuan": format_rupiah, "Total_Biaya": format_rupiah}), hide_index=True, use_container_width=True)
                
                # --- MESIN CETAK EXCEL (FIT TO 1 PAGE, COMPACT ROWS) ---
                def export_excel_rab(df_header, df_items):
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, Border, Side
                    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "RAB Export"
                    
                    ws.column_dimensions['A'].width = 15 # Kode
                    ws.column_dimensions['B'].width = 45 # Uraian
                    ws.column_dimensions['C'].width = 25 # Volume & Satuan
                    ws.column_dimensions['D'].width = 16 # Harga
                    ws.column_dimensions['E'].width = 16 # Total
                    
                    ws.sheet_properties.pageSetUpPr.fitToPage = True
                    ws.page_setup.fitToHeight = 1
                    ws.page_setup.fitToWidth = 1
                    
                    font_bold = Font(bold=True); font_header = Font(bold=True, size=11); align_center = Alignment(horizontal="center", vertical="center")
                    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    ws.merge_cells('A1:E1')
                    t_rab = df_header.get('Tahun', pd.Series(['2027'])).iloc[0]
                    if t_rab == "-": t_rab = "2027"
                    
                    ws['A1'] = f"RINCIAN ANGGARAN BIAYA (RAB) FAKULTAS ILMU BUDAYA\nTAHUN ANGGARAN {t_rab}"
                    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws['A1'].font = font_header
                    ws.row_dimensions[1].height = 40

                    meta_rows = [
                        ("Kementerian/ Lembaga:", "(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI"), 
                        ("Unit Eselon II/ Satker:", "(17) Dirjen Diktiristek / (677524) UNIVERSITAS MULAWARMAN"),
                        ("Kegiatan:", df_header['Kegiatan'].iloc[0]), ("Sasaran Kegiatan:", df_header['Sasaran'].iloc[0]), 
                        ("Klasifikasi Rincian Output:", df_header['KRO'].iloc[0]),
                        ("Volume:", df_header['Volume'].iloc[0]), ("Satuan Ukur:", df_header['Satuan'].iloc[0]), 
                        ("Alokasi Dana (Total Belanja):", f"Rp. {df_items['Total_Biaya'].sum():,.0f}".replace(',','.'))
                    ]
                    rp = 2
                    for label, val in meta_rows:
                        ws.cell(row=rp, column=1, value=label).font = font_bold
                        ws.cell(row=rp, column=2, value=val)
                        ws.merge_cells(start_row=rp, start_column=2, end_row=rp, end_column=5)
                        rp += 1

                    rp += 1
                    for col_idx, text in enumerate(["Kode", "Rincian Belanja", "Volume & Satuan", "Harga Satuan", "Jumlah Biaya"], start=1):
                        cell = ws.cell(row=rp, column=col_idx, value=text); cell.font = font_bold; cell.alignment = align_center; cell.border = border_all
                    rp += 1

                    # PRINT HIRARKI PADAT (TANPA BARIS KOSONG)
                    def print_row(kode, urai, vol, hrg, tot, is_bold=False):
                        nonlocal rp
                        ws.cell(row=rp, column=1, value=kode).border = border_all
                        ws.cell(row=rp, column=2, value=urai).border = border_all
                        ws.cell(row=rp, column=3, value=vol).border = border_all
                        ws.cell(row=rp, column=4, value=hrg).border = border_all
                        if hrg != "": ws.cell(row=rp, column=4).number_format = '#,##0'
                        ws.cell(row=rp, column=5, value=tot).border = border_all; ws.cell(row=rp, column=5).number_format = '#,##0'
                        if is_bold: 
                            for col in range(1,6): ws.cell(row=rp, column=col).font = Font(bold=True)
                        rp += 1

                    total_seluruh = df_items["Total_Biaya"].sum()
                    
                    for head_col, indent in [('RO', ""), ('Komponen', "  "), ('Sub_Komponen', "    ")]:
                        if df_header[head_col].iloc[0] and str(df_header[head_col].iloc[0]).strip() not in ["", "-", "Tidak Ada Sub-Komponen"]:
                            k_val, u_val = split_kode(df_header[head_col].iloc[0])
                            
                            if "." in k_val and len(k_val.split(".")) == 2 and len(k_val.split(".")[0]) == 3:
                                k1, k2 = k_val.split(".")
                                print_row(k1, f"{indent}{u_val}", "", "", total_seluruh, True)
                                print_row(k2, f"{indent}", "", "", total_seluruh, True)
                            else:
                                print_row(k_val, f"{indent}{u_val}", "", "", total_seluruh, True)

                    for akun, group in df_items.groupby("Akun_Belanja"):
                        k_ak, u_ak = split_kode(akun)
                        print_row(k_ak, f"      {u_ak}", "", "", group['Total_Biaya'].sum(), True)
                        for _, r in group.iterrows():
                            v_sat = get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2'])
                            print_row("", f"        - {r['Uraian']}", v_sat, r['Harga_Satuan'], r['Total_Biaya'])
                            
                    rp += 2
                    bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
                    try: 
                        tobj = datetime.strptime(df_header['Tgl_Cetak'].iloc[0], "%Y-%m-%d")
                        tgl_str = f"Samarinda, {tobj.day} {bulan_indo[tobj.month-1]} {tobj.year}"
                    except: tgl_str = f"Samarinda, {df_header['Tgl_Cetak'].iloc[0]}"
                    
                    ws.cell(row=rp, column=4, value=tgl_str)
                    ws.cell(row=rp+1, column=4, value=df_header['Jabatan'].iloc[0])
                    ws.cell(row=rp+5, column=4, value=df_header['Nama_Pejabat'].iloc[0]).font = Font(underline="single", bold=True)
                    ws.cell(row=rp+6, column=4, value=f"NIP. {df_header['NIP_Pejabat'].iloc[0]}")

                    output = BytesIO(); wb.save(output)
                    return output.getvalue()

                # --- MESIN CETAK PDF (AUTO-SCALE LANDSCAPE/PORTRAIT, COMPACT ROWS) ---
                def export_pdf_rab(df_header, df_items, orientasi):
                    total_seluruh = df_items["Total_Biaya"].sum()
                    t_rab = df_header.get('Tahun', pd.Series(['2027'])).iloc[0]
                    if t_rab == "-": t_rab = "2027"
                    
                    try: 
                        tobj = datetime.strptime(df_header['Tgl_Cetak'].iloc[0], "%Y-%m-%d")
                        bulan_indo = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
                        tgl_str = f"Samarinda, {tobj.day} {bulan_indo[tobj.month-1]} {tobj.year}"
                    except: tgl_str = f"Samarinda, {df_header['Tgl_Cetak'].iloc[0]}"
                    
                    page_rule = "A4 landscape" if orientasi == "Landscape" else "A4 portrait"
                    
                    html = f"""
                    <!DOCTYPE html>
                    <html><head><meta charset="utf-8">
                    <style>
                        @page {{ size: {page_rule}; margin: 10mm; }}
                        body {{ font-family: 'Arial', sans-serif; font-size: 8.5pt; line-height: 1.2; }}
                        .judul {{ text-align: center; font-weight: bold; font-size: 11pt; margin-bottom: 15px; }}
                        .tabel-meta td {{ padding: 1px 3px; font-size: 8.5pt; }}
                        .tabel-utama {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 8pt; }}
                        .tabel-utama th, .tabel-utama td {{ border: 1px solid black; padding: 4px; }}
                        .tabel-utama th {{ background-color: #f2f2f2; text-align: center; }}
                        .text-right {{ text-align: right; }} .text-center {{ text-align: center; }} .bold {{ font-weight: bold; }}
                        .ttd-box {{ width: 220px; float: right; text-align: left; margin-top: 20px; margin-right: 15px; page-break-inside: avoid; }}
                    </style></head><body>
                    <div class="judul">RINCIAN ANGGARAN BIAYA (RAB) FAKULTAS ILMU BUDAYA<br>TAHUN ANGGARAN {t_rab}</div>
                    <table class="tabel-meta">
                        <tr><td class="bold">Kementerian/ Lembaga</td><td>:</td><td>(023) KEMENTERIAN PENDIDIKAN TINGGI, SAINS DAN TEKNOLOGI</td></tr>
                        <tr><td class="bold">Unit Eselon II/ Satker</td><td>:</td><td>(17) Dirjen Diktiristek / (677524) UNIVERSITAS MULAWARMAN</td></tr>
                        <tr><td class="bold">Kegiatan</td><td>:</td><td>{df_header['Kegiatan'].iloc[0]}</td></tr>
                        <tr><td class="bold">Sasaran Kegiatan</td><td>:</td><td>{df_header['Sasaran'].iloc[0]}</td></tr>
                        <tr><td class="bold">Klasifikasi Rincian Output</td><td>:</td><td>{df_header['KRO'].iloc[0]}</td></tr>
                        <tr><td class="bold">Volume</td><td>:</td><td>{df_header['Volume'].iloc[0]}</td></tr>
                        <tr><td class="bold">Satuan Ukur</td><td>:</td><td>{df_header['Satuan'].iloc[0]}</td></tr>
                        <tr><td class="bold">Alokasi Dana (Total Belanja)</td><td>:</td><td>Rp. {format_rupiah(total_seluruh)}</td></tr>
                    </table>
                    <table class="tabel-utama">
                        <tr><th>Kode</th><th>Rincian Belanja</th><th>Volume & Satuan</th><th>Harga Satuan</th><th>Jumlah Biaya</th></tr>
                    """
                    for head_col, indent in [('RO', ""), ('Komponen', "  "), ('Sub_Komponen', "    ")]:
                        if df_header[head_col].iloc[0] and str(df_header[head_col].iloc[0]).strip() not in ["", "-", "Tidak Ada Sub-Komponen"]:
                            k, u = split_kode(df_header[head_col].iloc[0])
                            if "." in k and len(k.split(".")) == 2 and len(k.split(".")[0]) == 3:
                                k1, k2 = k.split(".")
                                html += f"<tr><td class='bold'>{k1}</td><td class='bold'>{indent}{u}</td><td></td><td></td><td class='bold text-right'>{format_rupiah(total_seluruh)}</td></tr>"
                                html += f"<tr><td class='bold'>{k2}</td><td class='bold'>{indent}</td><td></td><td></td><td class='bold text-right'>{format_rupiah(total_seluruh)}</td></tr>"
                            else:
                                html += f"<tr><td class='bold'>{k}</td><td class='bold'>{indent}{u}</td><td></td><td></td><td class='bold text-right'>{format_rupiah(total_seluruh)}</td></tr>"
                    
                    for akun, group_akun in df_items.groupby("Akun_Belanja"):
                        k_ak, u_ak = split_kode(akun)
                        html += f"<tr><td class='bold'>{k_ak}</td><td class='bold'>      {u_ak}</td><td></td><td></td><td class='bold text-right'>{format_rupiah(group_akun['Total_Biaya'].sum())}</td></tr>"
                        for _, r in group_akun.iterrows():
                            v_sat_str = get_vol_sat_combined(r['Vol_1'], r['Sat_1'], r['Vol_2'], r['Sat_2'])
                            html += f"<tr><td></td><td>        - {r['Uraian']}</td><td class='text-center'>{v_sat_str}</td><td class='text-right'>{format_rupiah(r['Harga_Satuan'])}</td><td class='text-right'>{format_rupiah(r['Total_Biaya'])}</td></tr>"
                    
                    html += f"""</table>
                    <div class="ttd-box">
                        {tgl_str}<br>{df_header['Jabatan'].iloc[0]}<br><br><br><br><br>
                        <b><u>{df_header['Nama_Pejabat'].iloc[0]}</u></b><br>NIP. {df_header['NIP_Pejabat'].iloc[0]}
                    </div>
                    </body></html>"""
                    return html

                st.markdown("#### 🖨️ Cetak & Unduh Dokumen RAB")
                orientasi_pdf = st.radio("Pilih Orientasi PDF:", ["Landscape", "Portrait"], horizontal=True)

                col_x1, col_x2 = st.columns([1, 4])
                with col_x1:
                    st.download_button("📥 Download Excel Resmi", data=export_excel_rab(head_terpilih, detail_terpilih), file_name=f"RAB_{tahun_rab}_{pilih_arsip}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
                    st.download_button("📑 PDF: Cetak RAB (Web)", data=export_pdf_rab(head_terpilih, detail_terpilih, orientasi_pdf).encode('utf-8'), file_name=f"Cetak_RAB_{tahun_rab}_{pilih_arsip}.html", mime="text/html", help="Tekan Ctrl+P di browser lalu pilih opsi 'Fit to Page'.")
                with col_x2:
                    if st.button("🗑️ Hapus Dokumen Ini"):
                        df_rab_utama = df_rab_utama[df_rab_utama["ID_RAB"] != pilih_arsip]
                        df_rab_detail = df_rab_detail[df_rab_detail["ID_RAB"] != pilih_arsip]
                        save_table(df_rab_utama, "rab_utama"); save_table(df_rab_detail, "rab_detail")
                        st.success("Terhapus!"); st.rerun()
